from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image as XLImage

import config
from automation_costos.recalculate import read_compras_workbook
from automation_costos.calculations import prepare_compras_dataframe
from automation_costos.ajustes_pagos import AJUSTES_COLUMNS, COLUMNAS_AJUSTE, aplicar_ajustes, fetch_pagos_ajustes
from automation_costos.utils import clean_code, clean_code_series, ensure_parent, make_folio_series, to_number


# Imagen corporativa: mismos colores/tipografia que el Compras (excel_exporter).
TITLE_FONT = Font(bold=True, size=14, color="000000")
HEADER_FILL = PatternFill("solid", fgColor="FF00FD28")  # verde, igual que el Compras
HEADER_FONT = Font(bold=True, color="FF000000")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def _add_logo(ws) -> None:
    """Inserta el logo de Soriana en A2, igual que en el Compras."""
    logo_path = config.BASE_DIR / "templates" / "Soriana-Logo.png"
    if not logo_path.exists():
        return
    logo = XLImage(str(logo_path))
    logo.width = 190
    logo.height = 46
    ws.add_image(logo, "A2")


CONSOLIDADO_COLUMNS = [
    "Proveedor",
    "Folio",
    "Tienda/CEDIS",
    "Negocio",
    "Fecha Pedido",
    "Pedido",
    "Fec Recibo",
    "Nota Entrada",
    "Factura",
    "Total Pagado",
    "Debio Pagar",
    "Diferencia",
    "Observaciones Auditor",
    "Fecha Pago",
    "Documento de Pago",
]

# Cuando el proveedor trae devoluciones (MR8M/KG-14) el Consolidado suma las columnas de ajuste.
CONSOLIDADO_COLUMNS_AJUSTES = CONSOLIDADO_COLUMNS + COLUMNAS_AJUSTE

DETALLE_COLUMNS = [
    "Folio_Pedido",
    "FecPed",
    "Folio_NE",
    "FecRecibo",
    "Factura",
    "Tienda/CEDIS",
    "Division",
    "Categoria",
    "GciaCateg",
    "Material",
    "CodBarra",
    "Descripcion",
    "FactEmpaqu",
    "CantRec",
    "CtoUnitario_sistema",
    "Costo Unitario correcto",
    "Porc_IEPS",
    "IEPS_Aud",
    "Porc_IVA",
    "IVA_Aud",
    "Debio Pagar correcto",
]


def write_validation_workbook(compras_path: Path, output_path: Path) -> Path:
    compras_df = prepare_compras_dataframe(read_compras_workbook(compras_path))
    return write_validation_from_dataframe(compras_df, output_path)


# Las únicas columnas del Compras que alimentan la Validación. Copiar solo estas en vez
# del DataFrame completo (105 columnas) es lo que hace viable a los proveedores de más de
# un millón de renglones.
_COLUMNAS_FUENTE = [
    "vndnbr", "vndname", "strnbr", "rcvnbr", "po_org", "podt", "ponbr", "rcvdt",
    "invnbr", "invnbr_ne", "paychkdt_ne", "paychknbr_ne",
    "debio_pagar_ne", "tot_pagado_ne", "dif_det_ne", "concepto",
    "nombre_division", "po_groupdescrip", "grupoarticulo", "cltstyle", "codbarra",
    "itmdesc", "fact_empaq", "can_rec", "ctouni", "ctonto_edi", "cto_aud",
    "prieps_edi", "ieps_aud", "poriva_edi", "iva_aud", "imp_aud",
]


def write_validation_from_dataframe(compras_df: pd.DataFrame, output_path: Path) -> Path:
    output_path = Path(output_path)
    ensure_parent(output_path)

    presentes = [column for column in _COLUMNAS_FUENTE if column in compras_df.columns]
    compras_df = compras_df[presentes].copy()
    compras_df["folio"] = make_folio_series(
        compras_df.get("strnbr", pd.Series(dtype=object)),
        compras_df.get("rcvnbr", pd.Series(dtype=object)),
    )

    consolidated = build_consolidado(compras_df)
    consolidated, ajustes, cols_cons = aplicar_ajustes_a_consolidado(compras_df, consolidated)
    detail = build_detalle(compras_df, consolidated["Folio"].tolist())

    wb = Workbook()
    resumen = wb.active
    resumen.title = "Resumen"
    ws_consolidado = wb.create_sheet("Consolidado")
    ws_ajustes = wb.create_sheet("Ajustes") if not ajustes.empty else None
    ws_detalle = wb.create_sheet("Detalle PAGOS")

    _write_resumen(resumen, compras_df, consolidated)
    _write_consolidado(ws_consolidado, compras_df, consolidated, cols_cons)
    if ws_ajustes is not None:
        _write_ajustes(ws_ajustes, compras_df, ajustes)
    _write_detalle(ws_detalle, compras_df, detail)

    wb.save(output_path)
    return output_path


def aplicar_ajustes_a_consolidado(
    compras_df: pd.DataFrame, consolidated: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Cruza las devoluciones MR8M/KG-14 del proveedor y las aplica al Consolidado.

    Devuelve `(consolidado, ajustes, columnas)`. Es **no-op** (comportamiento idéntico al
    anterior, 15 columnas) si el proveedor no tiene devoluciones o si la consulta a la BD
    falla: la Validación se sigue generando aunque no haya SQL disponible.
    """
    vacio = pd.DataFrame(columns=AJUSTES_COLUMNS)
    if consolidated.empty or compras_df.empty:
        return consolidated, vacio, CONSOLIDADO_COLUMNS
    vndnbr = clean_code(compras_df.iloc[0].get("vndnbr"))
    if not vndnbr:
        return consolidated, vacio, CONSOLIDADO_COLUMNS
    try:
        pagos = fetch_pagos_ajustes(vndnbr)
    except Exception as exc:  # noqa: BLE001 — sin BD/permisos seguimos sin ajustes
        print(f"[Ajustes pagos] {vndnbr}: omitido ({exc})", flush=True)
        return consolidated, vacio, CONSOLIDADO_COLUMNS

    threshold = config.VALIDATION_DIFFERENCE_THRESHOLD
    con, ajustes = aplicar_ajustes(consolidated, pagos, threshold=threshold)
    if ajustes.empty:
        return consolidated, ajustes, CONSOLIDADO_COLUMNS

    # Los renglones que la devolución dejó en ~0 desaparecen del Consolidado (quedan en la
    # hoja "Ajustes"); los parcialmente ajustados se quedan con su diferencia ya reducida.
    dif_aj = pd.to_numeric(con["Diferencia Ajustada"], errors="coerce").fillna(0.0)
    cancelado = (con["Ajuste Pagos"].abs() > 0.005) & (dif_aj <= threshold)
    con = con[~cancelado].reset_index(drop=True)
    return con, ajustes, CONSOLIDADO_COLUMNS_AJUSTES


def build_consolidado(df: pd.DataFrame) -> pd.DataFrame:
    threshold = config.VALIDATION_DIFFERENCE_THRESHOLD
    work = _con_folio(df)
    use_concept_filter = _has_audit_concepts(work)

    # Prefiltro vectorizado: el Consolidado solo lleva los folios que califican, así que
    # se descartan ANTES de agrupar. Recorrer los ~100k folios de un proveedor grande uno
    # por uno en Python para tirar el 99% era el paso lento de la Validación.
    work = work[_folios_que_califican(work, use_concept_filter, threshold)]

    sort_cols = [column for column in ["folio", "podt", "rcvdt", "invnbr"] if column in work.columns]
    if sort_cols:
        work = work.sort_values(sort_cols)

    rows = []
    for folio, group in work.groupby("folio", dropna=False):
        first = group.iloc[0]
        debio = _num(first.get("debio_pagar_ne"))
        total_pagado = _num(first.get("tot_pagado_ne"))
        diferencia = _num(first.get("dif_det_ne"))
        rows.append(
            {
                "Proveedor": clean_code(first.get("vndnbr")),
                "Folio": folio,
                "Tienda/CEDIS": clean_code(first.get("strnbr")),
                "Negocio": first.get("po_org") or "SORIANA",
                "Fecha Pedido": first.get("podt"),
                "Pedido": clean_code(first.get("ponbr")),
                "Fec Recibo": first.get("rcvdt"),
                "Nota Entrada": clean_code(first.get("rcvnbr")),
                "Factura": first.get("invnbr_ne") or first.get("invnbr"),
                "Total Pagado": total_pagado,
                "Debio Pagar": debio,
                "Diferencia": diferencia,
                "Observaciones Auditor": "",
                "Fecha Pago": first.get("paychkdt_ne"),
                "Documento de Pago": clean_code(first.get("paychknbr_ne")),
            }
        )
    return pd.DataFrame(rows, columns=CONSOLIDADO_COLUMNS)


def _folios_que_califican(work: pd.DataFrame, use_concept_filter: bool, threshold: float) -> pd.Series:
    """Máscara de renglones cuyo folio entra al Consolidado.

    Misma regla que evaluaba el bucle por grupo: con clasificación del auditor entra el
    folio que tenga al menos un renglón "dif costos"; si no, el que su diferencia (valor
    por folio, igual en todos sus renglones) supere el umbral.
    """
    if use_concept_filter:
        es_dif = (
            work["concepto"].fillna("").astype(str).str.lower().str.strip().eq("dif costos")
        )
        return work["folio"].isin(work.loc[es_dif, "folio"].unique())
    if "dif_det_ne" not in work.columns:
        return pd.Series(False, index=work.index)
    diferencia = pd.to_numeric(work["dif_det_ne"], errors="coerce").fillna(0.0)
    return diferencia > threshold


def _has_audit_concepts(df: pd.DataFrame) -> bool:
    if "concepto" not in df.columns:
        return False
    concepts = df["concepto"].fillna("").astype(str).str.lower().str.strip()
    meaningful = concepts[concepts.ne("") & concepts.ne("nan")]
    return meaningful.isin({"dif costos", "sin diferencia", "sobrepago", "faltante"}).any()


def _con_folio(df: pd.DataFrame) -> pd.DataFrame:
    """Devuelve el DataFrame con la columna `folio`, sin copiarlo si ya la trae."""
    if "folio" in df.columns:
        return df
    work = df.copy()
    work["folio"] = make_folio_series(work["strnbr"], work["rcvnbr"])
    return work


def build_detalle(df: pd.DataFrame, folios: list[str]) -> pd.DataFrame:
    work = _con_folio(df)
    if folios:
        work = work[work["folio"].isin(folios)]
    rows = []
    for _, item in work.iterrows():
        rows.append(
            {
                "Folio_Pedido": clean_code(item.get("ponbr")),
                "FecPed": item.get("podt"),
                "Folio_NE": clean_code(item.get("rcvnbr")),
                "FecRecibo": item.get("rcvdt"),
                "Factura": item.get("invnbr_ne") or item.get("invnbr"),
                "Tienda/CEDIS": clean_code(item.get("strnbr")),
                "Division": item.get("nombre_division"),
                "Categoria": item.get("po_groupdescrip"),
                "GciaCateg": item.get("grupoarticulo"),
                "Material": clean_code(item.get("cltstyle")),
                "CodBarra": clean_code(item.get("codbarra")),
                "Descripcion": item.get("itmdesc"),
                "FactEmpaqu": _num(item.get("fact_empaq")),
                "CantRec": _num(item.get("can_rec")),
                "CtoUnitario_sistema": _num(item.get("ctouni")),
                "Costo Unitario correcto": _num(item.get("cto_aud")),
                "Porc_IEPS": _num(item.get("prieps_edi")),
                "IEPS_Aud": _num(item.get("ieps_aud")),
                "Porc_IVA": _num(item.get("poriva_edi")),
                "IVA_Aud": _num(item.get("iva_aud")),
                "Debio Pagar correcto": _num(item.get("imp_aud")),
            }
        )
    return pd.DataFrame(rows, columns=DETALLE_COLUMNS)


def _write_resumen(ws, compras_df: pd.DataFrame, consolidated: pd.DataFrame) -> None:
    vendor = _vendor_label(compras_df)
    _add_logo(ws)
    for row, text in ((2, "Tiendas Soriana, S.A. de C.V."), (3, vendor), (4, "Validacion de Condiciones")):
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        cell = ws.cell(row, 4, text)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 21
    ws["C7"] = "Resultado Auditoria"
    ws["D7"] = "Observaciones Auditor"
    dif_col = "Diferencia Ajustada" if "Diferencia Ajustada" in consolidated.columns else "Diferencia"
    ws["C8"] = float(consolidated[dif_col].sum()) if not consolidated.empty else 0
    ws["D8"] = "Diferencia costos" if not consolidated.empty else ""
    for cell in ["C7", "D7"]:
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = Alignment(horizontal="center")
    ws["C8"].number_format = '#,##0.00'
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.sheet_view.showGridLines = False


def _write_consolidado(
    ws, compras_df: pd.DataFrame, consolidated: pd.DataFrame, columnas: list[str] = CONSOLIDADO_COLUMNS
) -> None:
    _write_sheet_title(ws, compras_df)
    # Totales SUBTOTAL sobre las columnas de importe presentes (por nombre, no por posición,
    # para que aguante las columnas extra de ajustes sin desalinearse).
    for offset, nombre in enumerate(columnas):
        if nombre in {"Debio Pagar", "Diferencia", "Diferencia Ajustada", "Monto No Compensado"}:
            letra = get_column_letter(2 + offset)
            celda = ws[f"{letra}7"]
            celda.value = f"=SUBTOTAL(109,{letra}9:{letra}1048576)"
            celda.fill = TOTAL_FILL
            celda.number_format = '#,##0.00'
    _write_table(ws, consolidated, columnas, start_row=8, start_col=2)
    money_cols = {
        i + 1 for i, c in enumerate(columnas)
        if c in {"Total Pagado", "Debio Pagar", "Diferencia", "Ajuste Pagos", "Diferencia Ajustada", "Monto No Compensado"}
    }
    date_cols = {i + 1 for i, c in enumerate(columnas) if c in {"Fecha Pedido", "Fec Recibo", "Fecha Pago"}}
    _format_validation_sheet(ws, start_col=2, num_cols=len(columnas), money_cols=money_cols, date_cols=date_cols)


def _write_ajustes(ws, compras_df: pd.DataFrame, ajustes: pd.DataFrame) -> None:
    """Hoja "Ajustes": bitácora de las devoluciones (MR8M/KG-14) que anularon diferencias."""
    _write_sheet_title(ws, compras_df)
    nota = ws.cell(6, 2, "Devoluciones de pago MR8M / KG-14 (compensadas y pendientes)")
    nota.font = Font(bold=True, italic=True)
    _write_table(ws, ajustes, AJUSTES_COLUMNS, start_row=8, start_col=2)
    money_cols = {
        i + 1 for i, c in enumerate(AJUSTES_COLUMNS)
        if c in {"Diferencia Original", "Ajuste Aplicado", "Monto No Compensado"}
    }
    date_cols = {i + 1 for i, c in enumerate(AJUSTES_COLUMNS) if c == "Fecha Pago"}
    _format_validation_sheet(ws, start_col=2, num_cols=len(AJUSTES_COLUMNS), money_cols=money_cols, date_cols=date_cols)


def _write_detalle(ws, compras_df: pd.DataFrame, detail: pd.DataFrame) -> None:
    _write_sheet_title(ws, compras_df)
    ws["V7"] = "=SUBTOTAL(109,V9:V1048576)"
    ws["V7"].fill = TOTAL_FILL
    ws["V7"].number_format = '#,##0.00'
    _write_table(ws, detail, DETALLE_COLUMNS, start_row=8, start_col=2)
    _format_validation_sheet(ws, start_col=2, num_cols=len(DETALLE_COLUMNS), money_cols={15, 16, 21}, date_cols={2, 4}, percent_cols={17, 18, 19, 20})


def _write_sheet_title(ws, compras_df: pd.DataFrame) -> None:
    _add_logo(ws)
    for row, text in ((2, "Tiendas Soriana, S.A. de C.V."), (3, _vendor_label(compras_df)), (4, "Validacion de Condiciones")):
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=11)
        cell = ws.cell(row, 4, text)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 21
    ws.sheet_view.showGridLines = False


def _write_table(ws, df: pd.DataFrame, columns: list[str], start_row: int, start_col: int) -> None:
    for idx, column in enumerate(columns, start=start_col):
        cell = ws.cell(start_row, idx, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Los datos se vuelcan en el ORDEN de `columns`, no en el del DataFrame: si el df trae las
    # columnas en otro orden (p. ej. `Diferencia Ajustada` calculada al final en aplicar_ajustes),
    # `reindex` las realinea con los encabezados. Sin esto los valores salen corridos.
    datos = df.reindex(columns=columns)
    for row_idx, row in enumerate(datos.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row_idx, col_idx, None if pd.isna(value) else value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    end_row = start_row + max(len(df), 1)
    end_col = start_col + len(columns) - 1
    ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _format_validation_sheet(
    ws,
    start_col: int,
    num_cols: int,
    money_cols: set[int],
    date_cols: set[int],
    percent_cols: set[int] | None = None,
) -> None:
    percent_cols = percent_cols or set()
    widths = [14, 20, 13, 14, 13, 14, 13, 13, 18, 14, 14, 14, 24, 13, 18]
    for offset in range(num_cols):
        col_idx = start_col + offset
        width = widths[offset] if offset < len(widths) else 14
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        data_col = offset + 1
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=9):
            for item in cell:
                if data_col in money_cols:
                    item.number_format = '#,##0.00'
                elif data_col in date_cols:
                    item.number_format = "yyyy-mm-dd"
                elif data_col in percent_cols:
                    item.number_format = "0.00%"
    ws.freeze_panes = "B9"


def _vendor_label(df: pd.DataFrame) -> str:
    if df.empty:
        return ""
    vendor = clean_code(df.iloc[0].get("vndnbr"))
    name = df.iloc[0].get("vndname") or ""
    return f"{vendor} - {name}".strip(" -")


def _num(value) -> float:
    parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(parsed):
        return 0.0
    return float(parsed)


# ---------------------------------------------------------------------------------------
# Validación ESCALABLE (para proveedores gigantes): mismo contenido, motor xlsxwriter en
# streaming + Detalle vectorizado. openpyxl no aguanta millones de renglones; esto sí.
# ---------------------------------------------------------------------------------------

def build_detalle_rapido(df: pd.DataFrame, folios: list[str]) -> pd.DataFrame:
    """Versión vectorizada de `build_detalle` (mismas columnas), sin el bucle `iterrows`.

    Con millones de renglones `iterrows` tarda horas y la lista de dicts agota la memoria;
    esto arma el Detalle con operaciones de columna."""
    work = _con_folio(df)
    if folios:
        work = work[work["folio"].isin(folios)]

    def col(nombre: str) -> pd.Series:
        return work[nombre] if nombre in work.columns else pd.Series(pd.NA, index=work.index)

    factura = col("invnbr_ne")
    vacia = factura.isna() | factura.astype(str).str.strip().isin(["", "None", "nan"])
    factura = factura.mask(vacia, col("invnbr"))

    out = pd.DataFrame(
        {
            "Folio_Pedido": clean_code_series(col("ponbr")),
            "FecPed": col("podt"),
            "Folio_NE": clean_code_series(col("rcvnbr")),
            "FecRecibo": col("rcvdt"),
            "Factura": factura,
            "Tienda/CEDIS": clean_code_series(col("strnbr")),
            "Division": col("nombre_division"),
            "Categoria": col("po_groupdescrip"),
            "GciaCateg": col("grupoarticulo"),
            "Material": clean_code_series(col("cltstyle")),
            "CodBarra": clean_code_series(col("codbarra")),
            "Descripcion": col("itmdesc"),
            "FactEmpaqu": to_number(col("fact_empaq")),
            "CantRec": to_number(col("can_rec")),
            "CtoUnitario_sistema": to_number(col("ctouni")),
            "Costo Unitario correcto": to_number(col("cto_aud")),
            "Porc_IEPS": to_number(col("prieps_edi")),
            "IEPS_Aud": to_number(col("ieps_aud")),
            "Porc_IVA": to_number(col("poriva_edi")),
            "IVA_Aud": to_number(col("iva_aud")),
            "Debio Pagar correcto": to_number(col("imp_aud")),
        }
    )
    return out[DETALLE_COLUMNS]


def write_validation_rapida(compras_df: pd.DataFrame, output_path: Path, *, en_sitio: bool = False) -> Path:
    """Escribe la Validación con xlsxwriter (streaming) — mismo contenido que
    `write_validation_from_dataframe`, pero aguanta millones de renglones sin agotar memoria.

    Reusa `build_consolidado` (folio-nivel, chico) y `build_detalle_rapido` (vectorizado).

    Con `en_sitio=True` **muta** el DataFrame recibido (recorta columnas y agrega `folio`)
    en vez de copiarlo: quien llama debe poseerlo y no volver a usarlo.

    Para proveedores gigantes cuyo detalle no cabe en memoria de golpe, ver
    `write_validation_streaming` (escribe la misma hoja Detalle por chunks).
    """
    import xlsxwriter

    output_path = Path(output_path)
    ensure_parent(output_path)

    presentes = [c for c in _COLUMNAS_FUENTE if c in compras_df.columns]
    if en_sitio:
        sobra = [c for c in compras_df.columns if c not in presentes and c != "folio"]
        if sobra:
            compras_df.drop(columns=sobra, inplace=True)
    else:
        compras_df = compras_df[presentes].copy()
    if "folio" not in compras_df.columns:
        compras_df["folio"] = make_folio_series(
            compras_df.get("strnbr", pd.Series(dtype=object)),
            compras_df.get("rcvnbr", pd.Series(dtype=object)),
        )

    consolidated = build_consolidado(compras_df)
    consolidated, ajustes, cols_cons = aplicar_ajustes_a_consolidado(compras_df, consolidated)
    detail = build_detalle_rapido(compras_df, consolidated["Folio"].tolist())
    # Ya no se necesita la tabla fuente; liberarla antes de la fase de escritura baja el pico.
    del compras_df

    wb = xlsxwriter.Workbook(str(output_path), {"constant_memory": True, "use_zip64": True})
    try:
        hdr = wb.add_format({"bold": True, "bg_color": "#00FD28", "border": 1})
        money = wb.add_format({"num_format": "#,##0.00"})

        resumen = wb.add_worksheet("Resumen")
        dif_col = "Diferencia Ajustada" if "Diferencia Ajustada" in consolidated.columns else "Diferencia"
        total = float(pd.to_numeric(consolidated[dif_col], errors="coerce").fillna(0).sum())
        resumen.write(6, 2, "Resultado Auditoria", hdr)
        resumen.write(6, 3, "Observaciones Auditor", hdr)
        resumen.write(7, 2, total, money)
        resumen.write(7, 3, "Diferencia costos" if not consolidated.empty else "")

        _dump_tabla(wb, "Consolidado", cols_cons, consolidated, hdr)
        if not ajustes.empty:
            _dump_tabla(wb, "Ajustes", AJUSTES_COLUMNS, ajustes, hdr)
        _dump_tabla(wb, "Detalle PAGOS", DETALLE_COLUMNS, detail, hdr)
        wb.close()
    except BaseException:
        try:
            wb.close()
        except BaseException:
            pass
        output_path.unlink(missing_ok=True)
        raise
    return output_path


# --- Formato para las hojas escritas con xlsxwriter (streaming, proveedores gigantes) ------
# Mismo lenguaje visual que la ruta openpyxl: logo, título, encabezado verde, formatos de
# número, fórmulas de total, freeze y autofiltro. La diferencia es solo el motor.

_XLSX_START_COL = 1  # columna B (0-indexed), igual que la ruta openpyxl

_XLSX_MONEY = {
    "Total Pagado", "Debio Pagar", "Diferencia", "Ajuste Pagos", "Diferencia Ajustada",
    "Monto No Compensado", "Diferencia Original", "Ajuste Aplicado",
    "CtoUnitario_sistema", "Costo Unitario correcto", "Debio Pagar correcto",
}
_XLSX_PERCENT = {"Porc_IEPS", "IEPS_Aud", "Porc_IVA", "IVA_Aud"}
_XLSX_TOTAL = {"Debio Pagar", "Diferencia", "Diferencia Ajustada", "Monto No Compensado", "Debio Pagar correcto"}
_XLSX_ANCHOS = {
    "Proveedor": 12, "Folio": 18, "Tienda/CEDIS": 13, "Negocio": 12, "Fecha Pedido": 13,
    "Pedido": 13, "Fec Recibo": 13, "Nota Entrada": 13, "Factura": 16, "Total Pagado": 14,
    "Debio Pagar": 14, "Diferencia": 14, "Observaciones Auditor": 24, "Fecha Pago": 13,
    "Documento de Pago": 17, "Tipo Ajuste": 12, "Ajuste Pagos": 14, "Diferencia Ajustada": 16,
    "Compensado": 22, "Conteo No Compensados": 14, "Monto No Compensado": 16,
    "Folio_Pedido": 14, "FecPed": 13, "Folio_NE": 13, "FecRecibo": 13, "Division": 18,
    "Categoria": 16, "GciaCateg": 14, "Material": 13, "CodBarra": 16, "Descripcion": 30,
    "FactEmpaqu": 11, "CantRec": 10, "Costo Unitario correcto": 16, "Debio Pagar correcto": 18,
}


def _formatos_xlsx(wb) -> dict:
    return {
        "titulo": wb.add_format({"bold": True, "font_size": 14, "align": "center"}),
        "hdr": wb.add_format({"bold": True, "bg_color": "#00FD28", "border": 1,
                              "align": "center", "valign": "vcenter", "text_wrap": True}),
        "money": wb.add_format({"num_format": "#,##0.00"}),
        "pct": wb.add_format({"num_format": "0.00%"}),
        "total": wb.add_format({"num_format": "#,##0.00", "bg_color": "#E2F0D9", "bold": True}),
    }


def _titulo_xlsx(ws, vendor_label: str, fmts: dict) -> None:
    """Logo + título (filas 2-4), igual que `_write_sheet_title` de la ruta openpyxl."""
    ws.hide_gridlines(2)
    logo = config.BASE_DIR / "templates" / "Soriana-Logo.png"
    if logo.exists():
        ws.insert_image(1, 0, str(logo), {"x_scale": 0.9, "y_scale": 0.9})
    for fila, texto in ((1, "Tiendas Soriana, S.A. de C.V."), (2, vendor_label), (3, "Validacion de Condiciones")):
        ws.merge_range(fila, 3, fila, 10, texto, fmts["titulo"])


def _encabezar_xlsx(ws, columnas: list[str], fmts: dict, *, con_totales: bool) -> None:
    """Totales (fila 7), encabezado verde (fila 8), formatos de columna, freeze y autofiltro.

    Debe llamarse DESPUÉS de `_titulo_xlsx` (filas 2-4) y ANTES de escribir datos (fila 9+):
    `constant_memory` exige escribir en orden de filas."""
    from xlsxwriter.utility import xl_col_to_name

    if con_totales:
        for i, nombre in enumerate(columnas):
            if nombre in _XLSX_TOTAL:
                col = _XLSX_START_COL + i
                letra = xl_col_to_name(col)
                ws.write_formula(6, col, f"=SUBTOTAL(109,{letra}9:{letra}1048576)", fmts["total"])
    for i, nombre in enumerate(columnas):
        ws.write(7, _XLSX_START_COL + i, nombre, fmts["hdr"])
    for i, nombre in enumerate(columnas):
        col = _XLSX_START_COL + i
        fmt = fmts["money"] if nombre in _XLSX_MONEY else fmts["pct"] if nombre in _XLSX_PERCENT else None
        ws.set_column(col, col, _XLSX_ANCHOS.get(nombre, 14), fmt)
    ws.freeze_panes(8, _XLSX_START_COL)
    ws.autofilter(7, _XLSX_START_COL, 7, _XLSX_START_COL + len(columnas) - 1)


def _volcar_xlsx(wb, nombre: str, columnas: list[str], df, vendor_label: str, fmts: dict, *, con_totales: bool) -> None:
    """Vuelca una tabla chica (Consolidado / Ajustes) formateada, empezando en la fila 9."""
    ws = wb.add_worksheet(nombre)
    _titulo_xlsx(ws, vendor_label, fmts)
    _encabezar_xlsx(ws, columnas, fmts, con_totales=con_totales)
    if df is None or df.empty:
        return
    datos = df if list(df.columns) == list(columnas) else df.reindex(columns=columnas)
    fila = 8
    for row in datos.itertuples(index=False, name=None):
        ws.write_row(fila, _XLSX_START_COL, [_limpiar(v) for v in row])
        fila += 1


def write_validation_streaming(consolidado_src, detalle_chunks, output_path: Path) -> Path:
    """Validación de un proveedor gigante con la hoja Detalle COMPLETA, sin agotar memoria.

    A diferencia de `write_validation_rapida`, nunca materializa los millones de renglones de
    detalle a la vez:

    - `consolidado_src`: un renglón por folio (con los importes por folio) → para el
      Consolidado y los Ajustes. Es chico (un renglón por folio).
    - `detalle_chunks`: iterable de DataFrames (p. ej. un trimestre cada uno). Cada chunk se
      transforma con `build_detalle_rapido` y se **escribe de inmediato** en la hoja
      "Detalle PAGOS"; solo un chunk vive en memoria a la vez.
    """
    import xlsxwriter

    output_path = Path(output_path)
    ensure_parent(output_path)

    consolidated = build_consolidado(consolidado_src)
    consolidated, ajustes, cols_cons = aplicar_ajustes_a_consolidado(consolidado_src, consolidated)
    folios = set(consolidated["Folio"].tolist())
    vendor_label = _vendor_label(consolidado_src)

    wb = xlsxwriter.Workbook(str(output_path), {"constant_memory": True, "use_zip64": True})
    try:
        fmts = _formatos_xlsx(wb)

        resumen = wb.add_worksheet("Resumen")
        _titulo_xlsx(resumen, vendor_label, fmts)
        dif_col = "Diferencia Ajustada" if "Diferencia Ajustada" in consolidated.columns else "Diferencia"
        total = float(pd.to_numeric(consolidated[dif_col], errors="coerce").fillna(0).sum())
        resumen.set_column(2, 2, 22)
        resumen.set_column(3, 3, 28)
        resumen.write(6, 2, "Resultado Auditoria", fmts["hdr"])
        resumen.write(6, 3, "Observaciones Auditor", fmts["hdr"])
        resumen.write_number(7, 2, total, fmts["money"])
        resumen.write(7, 3, "Diferencia costos" if not consolidated.empty else "")

        _volcar_xlsx(wb, "Consolidado", cols_cons, consolidated, vendor_label, fmts, con_totales=True)
        if not ajustes.empty:
            _volcar_xlsx(wb, "Ajustes", AJUSTES_COLUMNS, ajustes, vendor_label, fmts, con_totales=False)

        # Detalle en streaming: cada chunk se transforma y se escribe; se libera antes del siguiente.
        stream = _DetalleStream(wb, DETALLE_COLUMNS, vendor_label, fmts)
        for chunk in detalle_chunks:
            det = build_detalle_rapido(chunk, folios)
            stream.escribir(det)
            del chunk, det
        wb.close()
    except BaseException:
        try:
            wb.close()
        except BaseException:
            pass
        output_path.unlink(missing_ok=True)
        raise
    return output_path


_TOPE_FILAS_HOJA = 1_048_576  # tope duro de filas por hoja en Excel


class _DetalleStream:
    """Escribe la hoja 'Detalle PAGOS' por chunks, formateada, abriendo hojas de continuación
    al llegar al tope de filas de Excel."""

    def __init__(self, wb, columnas: list[str], vendor_label: str, fmts: dict) -> None:
        self.wb = wb
        self.columnas = list(columnas)
        self.vendor_label = vendor_label
        self.fmts = fmts
        self.parte = 0
        self.fila = 0
        self.ws = None
        self._nueva_hoja()

    def _nueva_hoja(self) -> None:
        nombre = "Detalle PAGOS" if self.parte == 0 else f"Detalle PAGOS ({self.parte + 1})"
        self.ws = self.wb.add_worksheet(nombre)
        _titulo_xlsx(self.ws, self.vendor_label, self.fmts)
        _encabezar_xlsx(self.ws, self.columnas, self.fmts, con_totales=True)
        self.fila = 8  # los datos empiezan en la fila 9 (0-indexed 8)

    def escribir(self, df) -> None:
        if df is None or df.empty:
            return
        datos = df if list(df.columns) == self.columnas else df.reindex(columns=self.columnas)
        for fila in datos.itertuples(index=False, name=None):
            if self.fila >= _TOPE_FILAS_HOJA:
                self.parte += 1
                self._nueva_hoja()
            self.ws.write_row(self.fila, _XLSX_START_COL, [_limpiar(v) for v in fila])
            self.fila += 1


def _dump_tabla(wb, nombre: str, columnas: list[str], df: pd.DataFrame, hdr) -> None:
    """Vuelca una tabla (encabezado + datos) a una o varias hojas con xlsxwriter.

    Si los datos pasan del tope de filas de Excel, se continúan en "<nombre> (2)", etc.,
    para no truncar en silencio (el Detalle de un proveedor grande puede tener millones)."""
    total = 0 if df is None else len(df)
    capacidad = _TOPE_FILAS_HOJA - 1  # una fila la ocupa el encabezado
    partes = max(1, math.ceil(total / capacidad)) if total else 1

    # Si el df ya trae las columnas en el orden pedido (caso del Detalle gigante), se itera
    # directo: `df[columnas]` construiría un frame nuevo consolidando bloques (una copia de
    # millones de renglones) y era justo lo que agotaba la memoria en Arca.
    if not total:
        filas = iter(())
    elif list(df.columns) == list(columnas):
        filas = df.itertuples(index=False, name=None)
    else:
        filas = df[columnas].itertuples(index=False, name=None)
    for parte in range(partes):
        hoja = nombre if parte == 0 else f"{nombre} ({parte + 1})"
        ws = wb.add_worksheet(hoja)
        for col_idx, columna in enumerate(columnas):
            ws.write(0, col_idx, columna, hdr)
        escritas = 0
        fila_idx = 1
        for fila in filas:
            ws.write_row(fila_idx, 0, [_limpiar(v) for v in fila])
            fila_idx += 1
            escritas += 1
            if escritas >= capacidad:
                break  # el resto va en la siguiente hoja


def _limpiar(value):
    if isinstance(value, str):
        return value
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return value
