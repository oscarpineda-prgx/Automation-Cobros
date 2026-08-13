"""Reporte consolidado de diferencias por proveedor, periodo, monto y concepto.

Lee las Validaciones de Condiciones ya generadas y las resume en UN archivo de control
(`Reporte_Diferencias_Consolidado.xlsx`) que vive en la raíz de los entregables, al lado
de las carpetas de proveedor. Es lo que pidió Héctor Saucedo el 2026-08-13: un solo lugar
donde ver la diferencia de cada proveedor, con periodo, monto y concepto.

De dónde sale cada número
-------------------------
La fuente es la hoja **Consolidado** de cada `Validacion_<prov>_<nombre>.xlsx`: un renglón
por nota de entrada que quedó con diferencia. No se recalcula nada — el reporte es un
espejo agregado del entregable, así que **siempre cuadra** con lo que se le mandó al
cliente. La hoja **Ajustes** (cuando existe) aporta lo compensado por devoluciones
MR8M/KG-14, que es informativo: no se reclama.

    Diferencia detectada = Diferencia a reclamar + Compensado por devoluciones

`Diferencia a reclamar` es exactamente el número de la hoja "Resumen" de la Validación.

Concepto
--------
Si el auditor escribió algo en `Observaciones Auditor`, **eso** es el concepto: el reporte
se afina solo conforme se trabajan los archivos. Si no, se deriva del dato:

- `Diferencia de costos` — el caso normal.
- `Diferencia de costos (parcialmente compensada)` — hubo devolución pero quedó saldo.

Actualización automática
------------------------
`actualizar_reporte()` se llama sola al terminar cada proveedor (pipeline y GUI) y al
cerrar un lote. Es **incremental**: cachea lo ya leído por (archivo, fecha, tamaño), así
que un lote nuevo solo paga la lectura de lo nuevo. Releer la Validación de Arca son 30 s;
sin caché el reporte tardaría minutos en cada corrida.

Solo se considera `Validacion_<carpeta>.xlsx` con el nombre EXACTO de su carpeta. Eso
descarta las variantes que conviven en la misma carpeta (`..._2020.xlsx` de Arca,
`Validacion_rejecutado_...`, `..._P.xlsx`), que duplicarían los montos.
"""

from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd

import config
from automation_costos.utils import ensure_parent

NOMBRE_REPORTE = "Reporte_Diferencias_Consolidado.xlsx"

CONCEPTO_BASE = "Diferencia de costos"
CONCEPTO_PARCIAL = "Diferencia de costos (parcialmente compensada)"

# Tolerancia para decidir si una devolución realmente tocó el folio (centavos).
_EPS = 0.005

COLUMNAS_DETALLE = [
    "Proveedor", "Nombre", "Año", "Trimestre", "Concepto", "Folio", "Tienda/CEDIS",
    "Nota Entrada", "Factura", "Fec Recibo", "Total Pagado", "Debio Pagar",
    "Diferencia", "Ajuste Pagos", "Diferencia a reclamar", "Fecha Pago",
    "Documento de Pago", "Observaciones Auditor",
]

COLUMNAS_PERIODO = [
    "Proveedor", "Nombre", "Año", "Trimestre", "Concepto", "Notas", "Diferencia a reclamar",
]

COLUMNAS_RESUMEN = [
    "Proveedor", "Nombre", "Periodo", "Años", "Notas con diferencia",
    "Total Pagado", "Debio Pagar", "Diferencia a reclamar",
    "Compensado por devoluciones", "Diferencia detectada", "Concepto",
    "Monto inicial", "Variación vs inicial", "Revisiones",
    "Generado", "Archivo",
]

# Bitácora de montos. El reporte lee el Excel de la Validación, así que refleja SIEMPRE el
# último estado: si el auditor revisa y ajusta, un proveedor de $6.0 M pasa a $5.6 M y el
# monto original se pierde. Este histórico lo conserva: se anota un renglón cada vez que las
# cifras de un proveedor cambian, así que en cualquier momento se puede decir cuánto dio la
# primera corrida, cuánto da hoy y qué se movió en medio. Es append-only: nunca se reescribe.
NOMBRE_HISTORICO = "Historico_Diferencias.parquet"

COLUMNAS_HISTORICO = [
    "Fecha", "Proveedor", "Nombre", "Notas con diferencia",
    "Diferencia a reclamar", "Compensado por devoluciones", "Diferencia detectada", "Archivo",
]

# Lo que se compara para decidir si hubo cambio digno de anotarse.
_VIGILADAS = ["Notas con diferencia", "Diferencia a reclamar", "Compensado por devoluciones"]

_MONEY = {
    "Total Pagado", "Debio Pagar", "Diferencia", "Ajuste Pagos",
    "Diferencia a reclamar", "Compensado por devoluciones", "Diferencia detectada",
    "Monto inicial", "Variación vs inicial",
}
_TOTALIZA = {
    "Total Pagado", "Debio Pagar", "Diferencia", "Diferencia a reclamar",
    "Compensado por devoluciones", "Diferencia detectada",
    "Monto inicial", "Variación vs inicial",
}
_ANCHOS = {
    "Proveedor": 11, "Nombre": 42, "Periodo": 20, "Años": 24, "Concepto": 38,
    "Folio": 19, "Factura": 18, "Observaciones Auditor": 26, "Archivo": 46,
    "Notas con diferencia": 12, "Notas": 9, "Año": 8, "Trimestre": 10,
    "Generado": 17, "Documento de Pago": 17, "Nota Entrada": 13, "Fec Recibo": 12,
    "Fecha Pago": 12, "Tienda/CEDIS": 12, "Fecha": 17, "Monto inicial": 16,
    "Variación vs inicial": 16, "Revisiones": 10,
}


# ---------------------------------------------------------------------------------------
# Localizar y leer las Validaciones
# ---------------------------------------------------------------------------------------

# Carpeta de reejecución: manda sobre las carpetas normales. Los entregables que están ahí
# se rehicieron YA con las devoluciones MR8M/KG-14; los de las carpetas de proveedor se
# generaron antes de esa regla y sobrestiman el monto a reclamar (Pepsico, $85 M de más).
# Se distingue por la hoja "Ajustes": las viejas no la tienen.
CARPETA_REEJECUCION = "Rejecución_validación_pagos"


def _reejecutadas(raiz: Path) -> dict[str, Path]:
    """`{carpeta_de_proveedor: Validación que manda}` según la carpeta de reejecución."""
    carpeta = raiz / CARPETA_REEJECUCION
    if not carpeta.is_dir():
        return {}
    return {
        archivo.stem.removeprefix("Validacion_"): archivo
        for archivo in sorted(carpeta.glob("Validacion_*.xlsx"))
        if not archivo.name.startswith("~$")
    }


def validaciones(raiz: Path) -> list[tuple[str, Path]]:
    """`(carpeta_de_proveedor, Validación)` para cada proveedor, sin duplicar montos.

    Prioridad:

    1. Si el proveedor está en `Rejecución_validación_pagos`, **esa manda**. Es la corrida
       buena, la que ya contempla MR8M/KG-14.
    2. Si no, la Validación completa `Validacion_<carpeta>.xlsx` de su carpeta. Esa gana
       sobre las variantes que conviven ahí: Arca guarda además `..._2020` … `_2025`
       (subconjuntos del mismo total) y Pepsico una reejecución parcial.
    3. Si tampoco existe la completa, todas las `Validacion_<carpeta>_*.xlsx`, que ahí sí
       son complementarias (3M estaba partido en `_2020-2024` y `_2025`).
    """
    reejecutadas = _reejecutadas(raiz)
    encontradas: list[tuple[str, Path]] = []
    for carpeta in sorted(p for p in raiz.iterdir() if p.is_dir()):
        if carpeta.name == CARPETA_REEJECUCION:
            continue
        if carpeta.name in reejecutadas:
            encontradas.append((carpeta.name, reejecutadas.pop(carpeta.name)))
            continue
        completa = carpeta / f"Validacion_{carpeta.name}.xlsx"
        if completa.exists():
            encontradas.append((carpeta.name, completa))
            continue
        encontradas.extend(
            (carpeta.name, v) for v in sorted(carpeta.glob(f"Validacion_{carpeta.name}_*.xlsx"))
        )
    # Reejecutados sin carpeta propia: no se pierden, se reportan igual.
    encontradas.extend((nombre, ruta) for nombre, ruta in sorted(reejecutadas.items()))
    return encontradas


def _partes_del_nombre(carpeta: str) -> tuple[str, str]:
    numero, _, nombre = carpeta.partition("_")
    return numero.strip(), nombre.strip()


def _leer_hoja(path: Path, hoja: str, columnas: Iterable[str]) -> pd.DataFrame:
    """Lee una hoja de la Validación con openpyxl en modo lectura, sin tocar las demás.

    La Validación de un proveedor gigante pesa cientos de MB por su hoja "Detalle PAGOS";
    `read_only` permite abrirla en décimas de segundo y recorrer solo la hoja pedida.
    El encabezado se **busca** en vez de asumir su fila: la ruta openpyxl y la ruta
    xlsxwriter escriben el título con distinto número de filas de adorno.
    """
    import openpyxl

    quiero = list(columnas)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if hoja not in wb.sheetnames:
            return pd.DataFrame(columns=quiero)
        ws = wb[hoja]
        posicion: dict[str, int] | None = None
        filas: list[list] = []
        for fila in ws.iter_rows(values_only=True):
            if posicion is None:
                etiquetas = {str(v).strip(): i for i, v in enumerate(fila) if v is not None}
                if "Folio" in etiquetas or "Nota Entrada" in etiquetas:
                    posicion = {c: etiquetas[c] for c in quiero if c in etiquetas}
                continue
            filas.append([fila[i] if i < len(fila) else None for i in posicion.values()])
        if posicion is None:
            return pd.DataFrame(columns=quiero)
        df = pd.DataFrame(filas, columns=list(posicion))
    finally:
        wb.close()
    return df.reindex(columns=quiero)


_COLS_CONSOLIDADO = [
    "Proveedor", "Folio", "Tienda/CEDIS", "Fec Recibo", "Nota Entrada", "Factura",
    "Total Pagado", "Debio Pagar", "Diferencia", "Observaciones Auditor",
    "Fecha Pago", "Documento de Pago", "Ajuste Pagos", "Diferencia Ajustada",
]


# Códigos que Excel entrega mezclados (unos como texto y otros como número según cómo se
# capturaron). Se uniforman a texto: si no, ni el Parquet de la caché ni el filtro de Excel
# los aguantan en la misma columna.
_COMO_TEXTO = ["Folio", "Tienda/CEDIS", "Nota Entrada", "Factura", "Documento de Pago"]


def leer_validacion(path: Path, carpeta: str | None = None) -> tuple[pd.DataFrame, dict]:
    """Un renglón por nota con diferencia + los datos de devoluciones del proveedor.

    `carpeta` es el nombre de la carpeta del proveedor, de donde salen número y nombre. Se
    pasa aparte porque los archivos reejecutados NO viven en la carpeta del proveedor sino
    en `Rejecución_validación_pagos`, y ahí `path.parent.name` no identifica a nadie.
    """
    numero, nombre = _partes_del_nombre(carpeta or path.parent.name)
    df = _leer_hoja(path, "Consolidado", _COLS_CONSOLIDADO)
    for columna in _COMO_TEXTO:
        df[columna] = df[columna].map(_texto)

    if df.empty:
        detalle = pd.DataFrame(columns=COLUMNAS_DETALLE)
    else:
        # `Diferencia Ajustada` solo existe cuando el proveedor trae devoluciones; donde no,
        # la diferencia bruta ya es la reclamable. Hay que preguntarlo ANTES de rellenar con
        # ceros, o una columna ausente se confundiría con una llena de ceros legítimos.
        con_devoluciones = df["Diferencia Ajustada"].notna().any()
        for columna in ("Total Pagado", "Debio Pagar", "Diferencia", "Ajuste Pagos", "Diferencia Ajustada"):
            df[columna] = pd.to_numeric(df[columna], errors="coerce").fillna(0.0)
        recibo = pd.to_datetime(df["Fec Recibo"], errors="coerce")
        reclamar = df["Diferencia Ajustada"] if con_devoluciones else df["Diferencia"]

        detalle = pd.DataFrame(
            {
                "Proveedor": numero,
                "Nombre": nombre,
                "Año": recibo.dt.year.astype("Int64"),
                "Trimestre": _trimestre(recibo),
                "Concepto": _concepto(df),
                "Folio": df["Folio"],
                "Tienda/CEDIS": df["Tienda/CEDIS"],
                "Nota Entrada": df["Nota Entrada"],
                "Factura": df["Factura"],
                "Fec Recibo": recibo,
                "Total Pagado": df["Total Pagado"],
                "Debio Pagar": df["Debio Pagar"],
                "Diferencia": df["Diferencia"],
                "Ajuste Pagos": df["Ajuste Pagos"],
                "Diferencia a reclamar": reclamar,
                "Fecha Pago": pd.to_datetime(df["Fecha Pago"], errors="coerce"),
                "Documento de Pago": df["Documento de Pago"],
                "Observaciones Auditor": df["Observaciones Auditor"],
            }
        )[COLUMNAS_DETALLE]

    ajustes = _leer_hoja(path, "Ajustes", ["Ajuste Aplicado", "Fecha Pago"])
    if ajustes.empty:
        return detalle, {"compensado": 0.0, "pago_min": pd.NaT, "pago_max": pd.NaT}
    pagos = pd.to_datetime(ajustes["Fecha Pago"], errors="coerce").dropna()
    return detalle, {
        "compensado": float(pd.to_numeric(ajustes["Ajuste Aplicado"], errors="coerce").abs().sum()),
        # Sirve para darle periodo a los proveedores cuya diferencia se compensó COMPLETA: no
        # tienen ninguna nota en el Consolidado, así que no hay fecha de recibo de dónde sacarlo.
        "pago_min": pagos.min() if len(pagos) else pd.NaT,
        "pago_max": pagos.max() if len(pagos) else pd.NaT,
    }


def _texto(valor) -> str | None:
    """Un código como texto, sin el `.0` que Excel le cuelga a los que leyó como número."""
    if valor is None or (not isinstance(valor, str) and pd.isna(valor)):
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _trimestre(recibo: pd.Series) -> pd.Series:
    """"2024-T3" por fecha de recibo; vacío donde no hay fecha."""
    etiqueta = recibo.dt.year.astype("Int64").astype(str) + "-T" + recibo.dt.quarter.astype("Int64").astype(str)
    return etiqueta.where(recibo.notna(), None)


# Errores de Excel que aparecen en `Observaciones Auditor` cuando el auditor dejó un BUSCARV
# roto (Celaya trae 26,319 `#N/A`). No son clasificación: se ignoran y se usa el derivado.
_BASURA = {"", "nan", "none", "#n/a", "#ref!", "#value!", "#div/0!", "#name?", "#null!", "#num!"}


def _concepto(df: pd.DataFrame) -> pd.Series:
    """La observación del auditor manda; si no la hay, se deriva del ajuste por devolución."""
    observado = df["Observaciones Auditor"].fillna("").astype(str).str.strip()
    derivado = pd.Series(CONCEPTO_BASE, index=df.index)
    derivado[df["Ajuste Pagos"].abs() > _EPS] = CONCEPTO_PARCIAL
    return observado.where(~observado.str.lower().isin(_BASURA), derivado)


# ---------------------------------------------------------------------------------------
# Caché incremental
# ---------------------------------------------------------------------------------------

def _ruta_cache(raiz: Path, que: str) -> Path:
    """Una caché por raíz de entregables, dentro del proyecto (no ensucia la entrega)."""
    firma = hashlib.sha1(str(raiz.resolve()).lower().encode("utf-8")).hexdigest()[:8]
    return Path(config.OUTPUT_DIR) / f"cache_reporte_{que}_{firma}.parquet"


def _leer_cache(path: Path, log: Callable[[str], None]) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_parquet(path)
    except Exception as exc:  # noqa: BLE001 — una caché corrupta no debe tumbar el reporte
        log(f"  caché ilegible ({exc}); se relee todo")
        return pd.DataFrame()


def _huella(path: Path) -> str:
    info = path.stat()
    return f"{int(info.st_mtime)}:{info.st_size}"


def _recolectar(
    raiz: Path, *, forzar: bool, log: Callable[[str], None]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Devuelve `(notas, catalogo)`.

    El catálogo lleva un renglón por Validación leída, tenga o no diferencias. Sin él, un
    proveedor cuyo Consolidado salió vacío desaparecería del reporte y parecería no
    ejecutado — que es justo lo contrario de lo que dice el dato: se revisó y salió limpio.
    """
    cache_notas = _ruta_cache(raiz, "notas")
    cache_catalogo = _ruta_cache(raiz, "catalogo")
    previo = pd.DataFrame() if forzar else _leer_cache(cache_notas, log)
    previo_cat = pd.DataFrame() if forzar else _leer_cache(cache_catalogo, log)

    cacheado = (
        dict(zip(previo_cat["_archivo"], previo_cat["_huella"]))
        if "_archivo" in previo_cat.columns
        else {}
    )

    bloques: list[pd.DataFrame] = []
    catalogo: list[dict] = []
    nuevos = reusados = 0
    for carpeta, path in validaciones(raiz):
        clave = str(path.relative_to(raiz))
        huella = _huella(path)
        if cacheado.get(clave) == huella:
            if "_archivo" in previo.columns:
                bloques.append(previo[previo["_archivo"] == clave])
            catalogo.extend(previo_cat[previo_cat["_archivo"] == clave].to_dict("records"))
            reusados += 1
            continue
        log(f"  leyendo {carpeta} ...")
        detalle, info = leer_validacion(path, carpeta)
        detalle["_archivo"] = clave
        if not detalle.empty:
            bloques.append(detalle)
        numero, nombre = _partes_del_nombre(carpeta)
        catalogo.append(
            {
                "_archivo": clave,
                "_huella": huella,
                "Proveedor": numero,
                "Nombre": nombre,
                "_compensado": info["compensado"],
                "_pago_min": info["pago_min"],
                "_pago_max": info["pago_max"],
                "_generado": datetime.fromtimestamp(path.stat().st_mtime),
            }
        )
        nuevos += 1

    log(f"  {nuevos} validaciones leídas, {reusados} reutilizadas de la caché")
    notas = pd.concat(bloques, ignore_index=True) if bloques else pd.DataFrame(columns=[*COLUMNAS_DETALLE, "_archivo"])
    cat = pd.DataFrame(catalogo)

    for destino, df in ((cache_notas, notas), (cache_catalogo, cat)):
        try:
            ensure_parent(destino)
            df.to_parquet(destino, index=False)
        except Exception as exc:  # noqa: BLE001 — sin caché el reporte sigue siendo correcto
            log(f"  no se pudo guardar la caché {destino.name} ({exc})")
    return notas, cat


# ---------------------------------------------------------------------------------------
# Agregados
# ---------------------------------------------------------------------------------------

SIN_DIFERENCIAS = "Sin diferencias"
# Un proveedor puede quedarse SIN una sola nota en el Consolidado y aun así haber tenido
# diferencias: si las devoluciones (MR8M/KG-14) las anularon completas, `aplicar_ajustes_a_
# consolidado` las saca y quedan en la hoja Ajustes. Etiquetarlos "Sin diferencias" se leería
# como "no se les encontró nada", que es falso — son cinco proveedores y $3.8 M compensados.
TODO_COMPENSADO = "Compensado por devoluciones (nada que reclamar)"


def _resumen(datos: pd.DataFrame, catalogo: pd.DataFrame) -> pd.DataFrame:
    """Un renglón por proveedor ejecutado, recorra el catálogo y no las notas."""
    if catalogo.empty:
        return pd.DataFrame(columns=COLUMNAS_RESUMEN)

    por_archivo = dict(tuple(datos.groupby("_archivo"))) if not datos.empty else {}
    agrupado = catalogo.groupby(["Proveedor", "Nombre"], sort=False)

    filas = []
    for (numero, nombre), meta in agrupado:
        # Un proveedor puede estar partido en varias Validaciones complementarias (3M);
        # aquí se suman todas las que le corresponden.
        partes = [por_archivo[a] for a in meta["_archivo"] if a in por_archivo]
        grupo = pd.concat(partes, ignore_index=True) if partes else None
        compensado = float(meta["_compensado"].sum())

        if grupo is None or grupo.empty:
            # Sin notas no hay fecha de recibo; el periodo se toma de los pagos de las
            # devoluciones, que es la única fecha que queda. Se marca en la propia etiqueta
            # del periodo para no mezclar bases sin avisar.
            desde, hasta = meta["_pago_min"].min(), meta["_pago_max"].max()
            hay_fechas = pd.notna(desde) and pd.notna(hasta)
            filas.append(
                {
                    "Proveedor": numero,
                    "Nombre": nombre,
                    "Periodo": f"{desde:%Y-%m-%d} a {hasta:%Y-%m-%d} (pago devolución)" if hay_fechas else "",
                    "Años": ", ".join(str(a) for a in range(desde.year, hasta.year + 1)) if hay_fechas else "",
                    "Notas con diferencia": 0, "Total Pagado": 0.0, "Debio Pagar": 0.0,
                    "Diferencia a reclamar": 0.0,
                    "Compensado por devoluciones": compensado,
                    "Diferencia detectada": compensado,
                    "Concepto": TODO_COMPENSADO if compensado > _EPS else SIN_DIFERENCIAS,
                    "Generado": meta["_generado"].max(),
                    "Archivo": " | ".join(meta["_archivo"]),
                }
            )
            continue

        anios = sorted({int(a) for a in grupo["Año"].dropna().unique()})
        recibos = grupo["Fec Recibo"].dropna()
        # El concepto del proveedor es el que concentra el dinero, no el más frecuente:
        # una sola nota grande pesa más para el control que veinte chicas.
        por_concepto = grupo.groupby("Concepto")["Diferencia a reclamar"].sum()
        reclamar = float(grupo["Diferencia a reclamar"].sum())
        filas.append(
            {
                "Proveedor": numero,
                "Nombre": nombre,
                "Periodo": f"{recibos.min():%Y-%m-%d} a {recibos.max():%Y-%m-%d}" if len(recibos) else "",
                "Años": ", ".join(str(a) for a in anios),
                "Notas con diferencia": len(grupo),
                "Total Pagado": float(grupo["Total Pagado"].sum()),
                "Debio Pagar": float(grupo["Debio Pagar"].sum()),
                "Diferencia a reclamar": reclamar,
                "Compensado por devoluciones": compensado,
                "Diferencia detectada": reclamar + compensado,
                "Concepto": por_concepto.idxmax() if not por_concepto.empty else "",
                "Generado": meta["_generado"].max(),
                "Archivo": " | ".join(meta["_archivo"]),
            }
        )
    resumen = pd.DataFrame(filas, columns=COLUMNAS_RESUMEN)
    return resumen.sort_values("Diferencia a reclamar", ascending=False, ignore_index=True)


def _registrar_historico(raiz: Path, resumen: pd.DataFrame, log: Callable[[str], None]) -> pd.DataFrame:
    """Anota en la bitácora los proveedores cuyas cifras cambiaron y devuelve el histórico.

    Solo escribe cuando algo se movió, así que correr el reporte diez veces sin tocar nada no
    ensucia la bitácora. Un fallo aquí no interrumpe el reporte: se avisa y se sigue.
    """
    ruta = raiz / NOMBRE_HISTORICO
    previo = pd.DataFrame(columns=COLUMNAS_HISTORICO)
    if ruta.exists():
        try:
            previo = pd.read_parquet(ruta)
        except Exception as exc:  # noqa: BLE001
            log(f"  histórico ilegible ({exc}); se empieza uno nuevo")

    ultimo = (
        previo.sort_values("Fecha").drop_duplicates("Proveedor", keep="last").set_index("Proveedor")
        if not previo.empty
        else pd.DataFrame()
    )

    ahora = datetime.now()
    nuevos = []
    for _, fila in resumen.iterrows():
        proveedor = fila["Proveedor"]
        if proveedor in ultimo.index:
            anterior = ultimo.loc[proveedor]
            if all(abs(float(anterior[c]) - float(fila[c])) < 0.01 for c in _VIGILADAS):
                continue
        nuevos.append({"Fecha": ahora, **{c: fila[c] for c in COLUMNAS_HISTORICO[1:]}})

    if not nuevos:
        return previo
    historico = pd.concat([previo, pd.DataFrame(nuevos)], ignore_index=True)[COLUMNAS_HISTORICO]
    try:
        historico.to_parquet(ruta, index=False)
        log(f"  histórico: {len(nuevos)} proveedores con cambios anotados")
    except Exception as exc:  # noqa: BLE001 — el reporte vale aunque no se pueda anotar
        log(f"  no se pudo guardar el histórico ({exc})")
    return historico


def _con_historico(resumen: pd.DataFrame, historico: pd.DataFrame) -> pd.DataFrame:
    """Agrega al resumen el monto de la PRIMERA corrida y cuánto se ha movido desde entonces."""
    resumen = resumen.copy()
    if historico.empty:
        resumen["Monto inicial"] = resumen["Diferencia a reclamar"]
        resumen["Variación vs inicial"] = 0.0
        resumen["Revisiones"] = 1
        return resumen

    orden = historico.sort_values("Fecha")
    inicial = orden.drop_duplicates("Proveedor", keep="first").set_index("Proveedor")["Diferencia a reclamar"]
    veces = orden.groupby("Proveedor").size()
    resumen["Monto inicial"] = resumen["Proveedor"].map(inicial).fillna(resumen["Diferencia a reclamar"])
    resumen["Variación vs inicial"] = resumen["Diferencia a reclamar"] - resumen["Monto inicial"]
    resumen["Revisiones"] = resumen["Proveedor"].map(veces).fillna(1).astype(int)
    return resumen


def _por_periodo(datos: pd.DataFrame) -> pd.DataFrame:
    if datos.empty:
        return pd.DataFrame(columns=COLUMNAS_PERIODO)
    agrupado = (
        datos.groupby(["Proveedor", "Nombre", "Año", "Trimestre", "Concepto"], dropna=False)
        .agg(Notas=("Folio", "size"), **{"Diferencia a reclamar": ("Diferencia a reclamar", "sum")})
        .reset_index()
    )
    return agrupado.sort_values(
        ["Proveedor", "Año", "Trimestre"], ignore_index=True
    )[COLUMNAS_PERIODO]


# ---------------------------------------------------------------------------------------
# Escritura (mismo lenguaje visual que la Validación: logo, verde PRGX, totales, filtros)
# ---------------------------------------------------------------------------------------

_FILA_ENCABEZADO = 7  # 0-indexed; los datos arrancan en la 8 (fila 9 de Excel)


def _escribir(
    path: Path,
    resumen: pd.DataFrame,
    periodo: pd.DataFrame,
    detalle: pd.DataFrame,
    historico: pd.DataFrame,
) -> Path:
    """Escribe a un temporal y lo mueve encima del destino.

    El reporte se actualiza solo tras cada proveedor, así que tarde o temprano corre mientras
    alguien lo tiene abierto en Excel. Escribir directo dejaría el archivo bueno destruido a
    medias; así, si el destino está bloqueado, lo que había sigue intacto y el mensaje dice
    qué hacer. La siguiente corrida lo pone al día sola.
    """
    import os
    import xlsxwriter

    ensure_parent(path)
    temporal = path.with_name(f"~{path.stem}.tmp{path.suffix}")
    wb = xlsxwriter.Workbook(str(temporal), {"constant_memory": True, "use_zip64": True})
    try:
        fmts = {
            "titulo": wb.add_format({"bold": True, "font_size": 14, "align": "center"}),
            "sub": wb.add_format({"italic": True, "font_color": "#595959", "align": "center"}),
            "hdr": wb.add_format({"bold": True, "bg_color": "#00FD28", "border": 1,
                                  "align": "center", "valign": "vcenter", "text_wrap": True}),
            "money": wb.add_format({"num_format": "#,##0.00"}),
            "fecha": wb.add_format({"num_format": "yyyy-mm-dd"}),
            "total": wb.add_format({"num_format": "#,##0.00", "bg_color": "#E2F0D9", "bold": True}),
        }
        sello = f"Actualizado {datetime.now():%Y-%m-%d %H:%M} · {len(resumen)} proveedores"
        _hoja(wb, "Resumen por proveedor", COLUMNAS_RESUMEN, resumen, fmts, sello)
        _hoja(wb, "Por periodo y concepto", COLUMNAS_PERIODO, periodo, fmts, sello)
        _hoja(wb, "Historico de montos", COLUMNAS_HISTORICO, historico.sort_values(["Proveedor", "Fecha"]) if not historico.empty else historico, fmts, sello)
        _hoja(wb, "Detalle notas", COLUMNAS_DETALLE, detalle, fmts, sello)
        wb.close()
    except BaseException:
        try:
            wb.close()
        except BaseException:
            pass
        temporal.unlink(missing_ok=True)
        raise

    try:
        os.replace(temporal, path)
    except PermissionError:
        temporal.unlink(missing_ok=True)
        raise PermissionError(
            f"{path.name} está abierto en Excel; ciérralo y vuelve a correrlo "
            "(el archivo anterior quedó intacto)"
        ) from None
    return path


def _hoja(wb, nombre: str, columnas: list[str], df: pd.DataFrame, fmts: dict, sello: str) -> None:
    from xlsxwriter.utility import xl_col_to_name

    ws = wb.add_worksheet(nombre)
    ws.hide_gridlines(2)
    logo = config.RESOURCE_DIR / "templates" / "Soriana-Logo.png"
    if logo.exists():
        ws.insert_image(1, 0, str(logo), {"x_scale": 0.9, "y_scale": 0.9})
    ws.merge_range(1, 3, 1, 10, "Tiendas Soriana, S.A. de C.V.", fmts["titulo"])
    ws.merge_range(2, 3, 2, 10, "Auditoria de Costos 2020-2025 · Control de diferencias", fmts["titulo"])
    ws.merge_range(3, 3, 3, 10, sello, fmts["sub"])

    for i, columna in enumerate(columnas):
        col = 1 + i
        if columna in _TOTALIZA:
            letra = xl_col_to_name(col)
            ws.write_formula(5, col, f"=SUBTOTAL(109,{letra}9:{letra}1048576)", fmts["total"])
        ws.write(_FILA_ENCABEZADO, col, columna, fmts["hdr"])
        fmt = fmts["money"] if columna in _MONEY else None
        ws.set_column(col, col, _ANCHOS.get(columna, 15), fmt)
    ws.freeze_panes(_FILA_ENCABEZADO + 1, 2)
    ws.autofilter(_FILA_ENCABEZADO, 1, _FILA_ENCABEZADO, len(columnas))

    if df.empty:
        return
    datos = df.reindex(columns=columnas)
    fila = _FILA_ENCABEZADO + 1
    for renglon in datos.itertuples(index=False, name=None):
        for i, valor in enumerate(renglon):
            if valor is None or (not isinstance(valor, str) and pd.isna(valor)):
                continue
            if isinstance(valor, (pd.Timestamp, datetime)):
                ws.write_datetime(fila, 1 + i, valor.to_pydatetime() if hasattr(valor, "to_pydatetime") else valor, fmts["fecha"])
            else:
                ws.write(fila, 1 + i, valor)
        fila += 1


# ---------------------------------------------------------------------------------------
# Entrada pública
# ---------------------------------------------------------------------------------------

def actualizar_desde_validacion(
    validacion: Path | str, *, log: Callable[[str], None] = print
) -> Path | None:
    """Refresca el reporte a partir de una Validación recién escrita.

    Es el enganche del flujo manual: el auditor corrige el Compras, regenera la Validación
    (`main.py validate` o el botón de la GUI) y los montos cambian. Sin esto el reporte se
    quedaría con las cifras de la primera corrida, que es cuando MENOS trabajadas están.

    La raíz de entregables se deduce de la ruta: una Validación vive en
    `<raiz>/<carpeta_proveedor>/Validacion_<carpeta_proveedor>*.xlsx`. Si el archivo no tiene
    esa forma no se toca nada y se devuelve None — así una validación suelta escrita en
    cualquier carpeta no siembra reportes donde no van.
    """
    validacion = Path(validacion).resolve()
    carpeta = validacion.parent
    if not validacion.stem.startswith(f"Validacion_{carpeta.name}"):
        return None
    try:
        return actualizar_reporte(carpeta.parent, log=lambda m: None)
    except Exception as exc:  # noqa: BLE001 — la Validación ya está escrita y es lo que importa
        log(f"No se pudo actualizar el reporte de diferencias: {exc}")
        return None


def actualizar_reporte(
    raiz: Path | str,
    *,
    salida: Path | str | None = None,
    forzar: bool = False,
    log: Callable[[str], None] = print,
) -> Path:
    """Regenera el reporte consolidado leyendo las Validaciones bajo `raiz`.

    Devuelve la ruta del Excel. Es idempotente y barata de repetir: lo ya leído sale de la
    caché, así que puede colgarse del final de cada proveedor sin costo perceptible.
    """
    raiz = Path(raiz)
    if not raiz.exists():
        raise FileNotFoundError(f"No existe la carpeta de entregables: {raiz}")
    destino = Path(salida) if salida else raiz / NOMBRE_REPORTE

    datos, catalogo = _recolectar(raiz, forzar=forzar, log=log)
    resumen = _resumen(datos, catalogo)
    historico = _registrar_historico(raiz, resumen, log)
    resumen = _con_historico(resumen, historico)
    periodo = _por_periodo(datos)
    detalle = datos.reindex(columns=COLUMNAS_DETALLE)
    _escribir(destino, resumen, periodo, detalle, historico)
    compensados = int((resumen["Concepto"] == TODO_COMPENSADO).sum()) if not resumen.empty else 0
    limpios = int((resumen["Concepto"] == SIN_DIFERENCIAS).sum()) if not resumen.empty else 0
    log(
        f"  {len(resumen)} proveedores ({compensados} todo compensado, {limpios} sin diferencias) · "
        f"{len(detalle):,} notas · ${resumen['Diferencia a reclamar'].sum():,.2f} a reclamar "
        f"(${resumen['Diferencia detectada'].sum():,.2f} detectados)"
    )
    return destino
