"""Exception Report — consolida las diferencias de varios bloques ya trabajados.

Lee la hoja **Consolidado** de cada archivo de Validación de Condiciones (un "bloque") y
produce un solo reporte con dos pestañas, con la misma imagen corporativa (logo Soriana +
encabezados verdes) que la Validación:

- **Consolidado de diferencias**: un renglón por bloque (proveedor · periodo) con su monto de
  diferencia y su total de facturas/pedidos con diferencia.
- **Detalle de diferencia**: sumatoria de diferencia por **pedido + factura** (con su conteo),
  consolidando todos los bloques en uno solo.

Se lee **por nombre de columna** (no por posición): los archivos viejos traen columnas extra
(p. ej. `Liga`) y no queremos romper por eso. La hoja Consolidado nunca se parte (es nivel
folio), así que basta leerla directo.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd

import config
from automation_costos.utils import ensure_parent

# El encabezado del Consolidado NO está en una fila fija: los archivos chicos (openpyxl)
# lo ponen en la fila 8 (con título y subtotal arriba) y los gigantes (xlsxwriter) en la
# fila 1. Se busca por nombre dentro de las primeras filas.
_MAX_FILAS_ENCABEZADO = 30
_HOJA_CONSOLIDADO = "Consolidado"

# Columnas que necesitamos del Consolidado; se buscan por nombre.
_NECESARIAS = ["Proveedor", "Pedido", "Factura", "Diferencia", "Nota Entrada", "Tienda/CEDIS"]

HEADER_COLOR = "#00FD28"  # verde, igual que la Validación y el Compras
TITULO = "Consolidado Diferencias Costos Por Proveedor"

CONS_COLUMNS = [
    "Proveedor",
    "Nombre",
    "Periodo",
    "Diferencia",
    "Facturas con diferencia",
    "Pedidos con diferencia",
    "Folios con diferencia",
]
DET_COLUMNS = [
    "Proveedor",
    "Nombre",
    "Periodo",
    "Nota Entrada",
    "Tienda",
    "Pedido",
    "Factura",
    "Diferencia",
    "Registros",
]


def descubrir_validaciones(carpeta: Path | str) -> list[Path]:
    """Todos los `Validacion_*.xlsx` bajo `carpeta` (recursivo), ordenados por nombre."""
    carpeta = Path(carpeta)
    archivos = [
        p for p in carpeta.rglob("Validacion_*.xlsx")
        if not p.name.startswith("~$")  # ignora temporales de Excel abiertos
    ]
    return sorted(archivos, key=lambda p: p.name.lower())


def _partes_nombre(path: Path) -> tuple[str, str, str]:
    """De `Validacion_<num>_<nombre>[_<periodo>].xlsx` saca (numero, nombre, periodo)."""
    stem = path.stem
    resto = stem[len("Validacion_"):] if stem.startswith("Validacion_") else stem
    periodo = "2020-2025"
    match = re.search(r"_(\d{4}(?:-\d{4})?)$", resto)
    if match:
        periodo = match.group(1)
        resto = resto[: match.start()]
    numero, _, nombre = resto.partition("_")
    return numero.strip(), nombre.strip(), periodo


def leer_consolidado(path: Path) -> pd.DataFrame:
    """Lee la hoja Consolidado de un bloque y devuelve las columnas normalizadas.

    Columnas de salida: Numero, Nombre, Periodo, Pedido, Factura, Diferencia. Se leen TODOS
    los folios con diferencia (base 'todas', decisión de Óscar 2026-07-30); el estado de
    filas ocultas del auditor no se respeta aquí.

    Streaming con openpyxl `read_only` (no pandas): los Consolidados de Arca/Pepsico tienen
    ~150k renglones y `pd.read_excel` los carga enteros, lentísimo.
    """
    from openpyxl import load_workbook

    path = Path(path)
    numero_fn, nombre_fn, periodo = _partes_nombre(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if _HOJA_CONSOLIDADO not in wb.sheetnames:
            raise ValueError(f"'{path.name}' no tiene hoja '{_HOJA_CONSOLIDADO}'.")
        ws = wb[_HOJA_CONSOLIDADO]

        idx: dict[str, int] = {}
        header_found = False
        pedidos: list[str] = []
        facturas: list[str] = []
        notas: list[str] = []
        tiendas: list[str] = []
        difs: list[float] = []
        numero = numero_fn

        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # El encabezado no está en una fila fija: los archivos chicos (openpyxl) lo
            # ponen en la fila 8 y los gigantes (xlsxwriter) en la fila 1. Se busca la
            # primera fila que traiga TODAS las columnas necesarias por nombre.
            if not header_found:
                headers = [str(v).strip() if v is not None else "" for v in row]
                if all(name in headers for name in _NECESARIAS):
                    idx = {name: headers.index(name) for name in _NECESARIAS}
                    header_found = True
                elif r > _MAX_FILAS_ENCABEZADO:
                    raise ValueError(
                        f"'{path.name}' hoja Consolidado sin el encabezado esperado "
                        f"(columnas {_NECESARIAS}) en las primeras {_MAX_FILAS_ENCABEZADO} filas."
                    )
                continue

            prov = row[idx["Proveedor"]] if idx["Proveedor"] < len(row) else None
            dif = row[idx["Diferencia"]] if idx["Diferencia"] < len(row) else None
            if prov is None or str(prov).strip() == "":
                continue
            try:
                dif_num = float(dif)
            except (TypeError, ValueError):
                continue
            if numero == numero_fn:
                numero = str(prov).strip()
            ped = row[idx["Pedido"]] if idx["Pedido"] < len(row) else None
            fac = row[idx["Factura"]] if idx["Factura"] < len(row) else None
            nota = row[idx["Nota Entrada"]] if idx["Nota Entrada"] < len(row) else None
            tienda = row[idx["Tienda/CEDIS"]] if idx["Tienda/CEDIS"] < len(row) else None
            pedidos.append("" if ped is None else str(ped).strip())
            facturas.append("" if fac is None else str(fac).strip())
            notas.append("" if nota is None else str(nota).strip())
            tiendas.append("" if tienda is None else str(tienda).strip())
            difs.append(dif_num)
    finally:
        wb.close()

    return pd.DataFrame(
        {
            "Numero": numero,
            "Nombre": nombre_fn,
            "Periodo": periodo,
            "NotaEntrada": notas,
            "Tienda": tiendas,
            "Pedido": pedidos,
            "Factura": facturas,
            "Diferencia": difs,
        }
    )


def _span(periodos: Iterable[str]) -> str:
    """Combina varios periodos de un mismo proveedor en un solo rango (min-max).

    Así 3M, que se trabajó en dos bloques (2020-2024 y 2025), queda como un solo
    renglón con periodo '2020-2025'.
    """
    anios: list[int] = []
    for p in periodos:
        match = re.match(r"(\d{4})(?:-(\d{4}))?$", str(p).strip())
        if not match:
            continue
        anios.append(int(match.group(1)))
        anios.append(int(match.group(2) or match.group(1)))
    if not anios:
        return ""
    lo, hi = min(anios), max(anios)
    return str(lo) if lo == hi else f"{lo}-{hi}"


def construir(fuente: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Construye el Exception Report.

    Consolidado:
        Un renglón por proveedor.

    Detalle:
        Agrupa por Nota Entrada + Tienda.
        Pedido y Factura únicamente quedan como referencia.
    """

    fuente = fuente.copy()

    fuente["Periodo"] = (
        fuente.groupby("Numero")["Periodo"]
        .transform(lambda s: _span(s))
    )

    # ==========================================================
    # CONSOLIDADO
    # ==========================================================

    grupo_bloque = fuente.groupby(
        ["Numero", "Nombre", "Periodo"],
        dropna=False,
    )

    consolidado = (
        grupo_bloque
        .agg(
            Diferencia=("Diferencia", "sum"),
            Facturas=("Factura", "nunique"),
            Pedidos=("Pedido", "nunique"),
            Folios=("Diferencia", "size"),
        )
        .reset_index()
    )

    consolidado = consolidado.sort_values(
        "Diferencia",
        ascending=False
    ).reset_index(drop=True)

    consolidado.columns = CONS_COLUMNS

    # ==========================================================
    # DETALLE
    # AHORA LA LLAVE ES:
    # Nota Entrada + Tienda
    # ==========================================================

    detalle = (
        fuente.groupby(
            [
                "Numero",
                "Nombre",
                "Periodo",
                "NotaEntrada",
                "Tienda",
            ],
            dropna=False,
        )
        .agg(
            Pedido=(
                "Pedido",
                lambda x: ", ".join(
                    sorted(
                        {
                            str(v).strip()
                            for v in x
                            if pd.notna(v) and str(v).strip() != ""
                        }
                    )
                ),
            ),
            Factura=(
                "Factura",
                lambda x: ", ".join(
                    sorted(
                        {
                            str(v).strip()
                            for v in x
                            if pd.notna(v) and str(v).strip() != ""
                        }
                    )
                ),
            ),
            Diferencia=("Diferencia", "sum"),
            Registros=("Diferencia", "size"),
        )
        .reset_index()
    )

    detalle = detalle[
        [
            "Numero",
            "Nombre",
            "Periodo",
            "NotaEntrada",
            "Tienda",
            "Pedido",
            "Factura",
            "Diferencia",
            "Registros",
        ]
    ]

    detalle = detalle.sort_values(
        ["Numero", "Diferencia"],
        ascending=[True, False],
    ).reset_index(drop=True)

    detalle.columns = DET_COLUMNS

    return consolidado, detalle


def write_exception_report(
    entradas: Iterable[Path | str] | Path | str,
    output_path: Path | str,
    *,
    log: Callable[[str], None] = print,
) -> Path:
    """Genera el Exception Report a partir de archivos o de una carpeta a escanear."""
    paths = _resolver_entradas(entradas)
    if not paths:
        raise FileNotFoundError("No se encontraron archivos Validacion_*.xlsx para consolidar.")

    # Ignora archivos duplicados: si dos coinciden en (numero, periodo) es una copia
    # (p. ej. Validacion_741_..._P.xlsx). 3M sobrevive porque sus dos bloques tienen
    # periodo distinto (2020-2024 y 2025).
    vistos: set[tuple[str, str]] = set()
    unicos: list[Path] = []
    for path in paths:
        numero, _, periodo = _partes_nombre(Path(path))
        clave = (numero, periodo)
        if clave in vistos:
            log(f"  [=] Se ignora duplicado: {Path(path).name}")
            continue
        vistos.add(clave)
        unicos.append(Path(path))
    paths = unicos

    bloques: list[pd.DataFrame] = []
    for path in paths:
        try:
            df = leer_consolidado(path)
        except Exception as exc:  # noqa: BLE001 — un bloque ilegible no debe tumbar el reporte
            log(f"  [!] Se omite '{Path(path).name}': {exc}")
            continue
        log(f"  {Path(path).name}: {len(df):,} folios con diferencia · dif ${df['Diferencia'].sum():,.2f}")
        bloques.append(df)

    if not bloques:
        raise ValueError("Ningún bloque se pudo leer; no hay nada que consolidar.")

    fuente = pd.concat(bloques, ignore_index=True)
    consolidado, detalle = construir(fuente)
    log(
        f"  Total: {len(consolidado)} bloques · {len(detalle):,} pedido-factura · "
        f"dif ${consolidado['Diferencia'].sum():,.2f}"
    )

    output_path = Path(output_path)
    ensure_parent(output_path)
    _escribir_excel(consolidado, detalle, output_path)
    return output_path


def _resolver_entradas(entradas: Iterable[Path | str] | Path | str) -> list[Path]:
    if isinstance(entradas, (str, Path)):
        p = Path(entradas)
        return descubrir_validaciones(p) if p.is_dir() else [p]
    resueltas: list[Path] = []
    for entrada in entradas:
        p = Path(entrada)
        resueltas.extend(descubrir_validaciones(p) if p.is_dir() else [p])
    return resueltas


# ---------------------------------------------------------------------------------------
# Escritura con imagen corporativa (xlsxwriter: rápido y aguanta el Detalle grande)
# ---------------------------------------------------------------------------------------

def _escribir_excel(consolidado: pd.DataFrame, detalle: pd.DataFrame, output_path: Path) -> None:
    import xlsxwriter

    wb = xlsxwriter.Workbook(str(output_path), {"constant_memory": True, "use_zip64": True})
    try:
        fmt = {
            "titulo": wb.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter"}),
            "hdr": wb.add_format({"bold": True, "bg_color": HEADER_COLOR, "border": 1,
                                   "align": "center", "valign": "vcenter", "text_wrap": True}),
            "money": wb.add_format({"num_format": "#,##0.00"}),
            "int": wb.add_format({"num_format": "#,##0"}),
            "tot_lbl": wb.add_format({"bold": True, "bg_color": "#E2F0D9"}),
            "tot_money": wb.add_format({"bold": True, "bg_color": "#E2F0D9", "num_format": "#,##0.00"}),
            "tot_int": wb.add_format({"bold": True, "bg_color": "#E2F0D9", "num_format": "#,##0"}),
        }
        _hoja_consolidado(wb, fmt, consolidado)
        _hoja_detalle(wb, fmt, detalle)
        wb.close()
    except BaseException:
        try:
            wb.close()
        except BaseException:
            pass
        output_path.unlink(missing_ok=True)
        raise


def _encabezado_hoja(wb, ws, fmt, subtitulo: str, ncols: int) -> None:
    """Logo + título corporativo en las primeras filas, igual que la Validación."""
    ws.hide_gridlines(2)
    logo = config.RESOURCE_DIR / "templates" / "Soriana-Logo.png"
    if logo.exists():
        try:
            from PIL import Image as PILImage

            ancho = PILImage.open(logo).size[0]
            escala = 190.0 / ancho if ancho else 0.5
        except Exception:
            escala = 0.5
        ws.insert_image(1, 0, str(logo), {"x_scale": escala, "y_scale": escala})

    ultima = max(4, ncols - 1)
    for fila, texto in ((1, "Tiendas Soriana, S.A. de C.V."), (2, TITULO), (3, subtitulo)):
        ws.merge_range(fila, 3, fila, ultima, texto, fmt["titulo"])


def _suma_arriba(ws, fila_tot: int, fila0: int, col: int, n: int, valor, fmt) -> None:
    """Escribe en la fila de totales una fórmula =SUMA sobre las filas de datos de esa columna.

    Es fórmula viva (Excel la recalcula si editas o agregas datos dentro del rango) y guarda el
    valor ya calculado como caché para que se vea correcto aun antes de recalcular. El rango va
    de la primera a la última fila con datos.
    """
    from xlsxwriter.utility import xl_range

    if n <= 0:
        ws.write_number(fila_tot, col, 0, fmt)
        return
    rango = xl_range(fila0, col, fila0 + n - 1, col)  # p. ej. D7:D163
    ws.write_formula(fila_tot, col, f"=SUM({rango})", fmt, valor)


def _hoja_consolidado(wb, fmt, df: pd.DataFrame) -> None:
    ws = wb.add_worksheet("Consolidado de diferencias")
    _encabezado_hoja(wb, ws, fmt, "Consolidado de diferencias por proveedor", len(CONS_COLUMNS))

    # Sumatoria ARRIBA (fila 5): siempre visible aunque el archivo crezca; luego el
    # encabezado (fila 6) y los datos (fila 7+). Se congela para que no se pierdan al scrollear.
    fila_tot, fila_hdr, fila0 = 4, 5, 6
    n = len(df)
    ws.write(fila_tot, 0, "TOTAL", fmt["tot_lbl"])
    ws.write(fila_tot, 1, "", fmt["tot_lbl"])
    ws.write(fila_tot, 2, "", fmt["tot_lbl"])
    _suma_arriba(ws, fila_tot, fila0, 3, n, float(df["Diferencia"].sum()), fmt["tot_money"])
    _suma_arriba(ws, fila_tot, fila0, 4, n, int(df["Facturas con diferencia"].sum()), fmt["tot_int"])
    _suma_arriba(ws, fila_tot, fila0, 5, n, int(df["Pedidos con diferencia"].sum()), fmt["tot_int"])
    _suma_arriba(ws, fila_tot, fila0, 6, n, int(df["Folios con diferencia"].sum()), fmt["tot_int"])

    for c, nombre in enumerate(CONS_COLUMNS):
        ws.write(fila_hdr, c, nombre, fmt["hdr"])
    for i, row in enumerate(df.itertuples(index=False), start=fila0):
        ws.write(i, 0, row.Proveedor)
        ws.write(i, 1, row.Nombre)
        ws.write(i, 2, row.Periodo)
        ws.write_number(i, 3, float(row.Diferencia), fmt["money"])
        ws.write_number(i, 4, int(row._4), fmt["int"])
        ws.write_number(i, 5, int(row._5), fmt["int"])
        ws.write_number(i, 6, int(row._6), fmt["int"])

    ws.set_column(0, 0, 12)
    ws.set_column(1, 1, 42)
    ws.set_column(2, 2, 12)
    ws.set_column(3, 3, 18)
    ws.set_column(4, 6, 16)
    ws.autofilter(fila_hdr, 0, fila_hdr + len(df), len(CONS_COLUMNS) - 1)
    ws.freeze_panes(fila0, 0)


def _hoja_detalle(wb, fmt, df: pd.DataFrame) -> None:
    ws = wb.add_worksheet("Detalle de diferencia")
    _encabezado_hoja(wb, ws, fmt, "Sumatoria de diferencia por pedido y factura", len(DET_COLUMNS))

    # Sumatoria ARRIBA (fila 5), luego encabezado (fila 6) y datos (fila 7+).
    fila_tot, fila_hdr, fila0 = 4, 5, 6
    n = len(df)
    ws.write(fila_tot, 0, "TOTAL", fmt["tot_lbl"])
    for c in range(1, 7):
        ws.write(fila_tot, c, "", fmt["tot_lbl"])
    _suma_arriba(ws, fila_tot, fila0, 7, n, float(df["Diferencia"].sum()), fmt["tot_money"])
    _suma_arriba(ws, fila_tot, fila0, 8, n, int(df["Registros"].sum()), fmt["tot_int"])

    for c, nombre in enumerate(DET_COLUMNS):
        ws.write(fila_hdr, c, nombre, fmt["hdr"])
    for i, row in enumerate(df.itertuples(index=False), start=fila0):
        ws.write(i, 0, row.Proveedor)
        ws.write(i, 1, row.Nombre)
        ws.write(i, 2, row.Periodo)
        ws.write(i, 3, row._3)        # Nota Entrada
        ws.write(i, 4, row.Tienda)
        ws.write(i, 5, row.Pedido)
        ws.write(i, 6, row.Factura)
        ws.write_number(i, 7, float(row.Diferencia), fmt["money"])
        ws.write_number(i, 8, int(row.Registros), fmt["int"])

    ws.set_column(0, 0, 12)
    ws.set_column(1, 1, 40)
    ws.set_column(2, 2, 12)
    ws.set_column(3, 4, 18)
    ws.set_column(5, 6, 30)
    ws.set_column(7, 7, 18)
    ws.set_column(8, 8, 12)
    ws.autofilter(fila_hdr, 0, fila_hdr + len(df), len(DET_COLUMNS) - 1)
    ws.freeze_panes(fila0, 0)
