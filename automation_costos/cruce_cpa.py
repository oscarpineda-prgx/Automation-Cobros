"""Cruce del dataset de CPA Vision contra el archivo de Compras.

Llena el bloque de columnas EDI (`canfac_edi` .. `uuid`) de Compras con los datos del
CFDI descargados de CPA Vision.

La especificacion completa, con el origen de cada columna y la evidencia que respalda
cada decision, esta en `docs/MAPEO_CRUCE_CPA_COMPRAS.md`.

Resumen:

- Llave: RFC del proveedor + numero de factura + codigo de barras.
  `invnbr` de Compras trae la serie pegada y con formato inconsistente
  (`FN-21226`, `FN21226`, `-21226`), asi que se normaliza antes de cruzar.
- Ocho columnas se copian tal cual del CFDI, incluidos los importes de impuesto:
  medido sobre 66M de filas, `IVA`/`IEPS` coinciden al 100% con `Importe Impuesto`,
  mientras que despejarlos desde `Total` solo acierta en el 0.9% de los renglones.
- Dos columnas se calculan (`ctobto_edi`, `impart_edi`).
- `factem_edi` ya viene en Compras; no se toca.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from automation_costos.utils import to_number

# Columnas de Compras que se copian tal cual desde CPA Vision.
COLUMNAS_COPIADAS: dict[str, str] = {
    "canfac_edi": "Cantidad",
    "ctonto_edi": "Valor Unitario",
    "prieps_edi": "TASA IEPS",
    "imieps_edi": "IEPS",
    "poriva_edi": "TASA IVA",
    "impiva_edi": "IVA",
    "totfactura": "Total",
    "uuid": "UUID",
}

# `factem_edi` no viene del CFDI: se toma de una columna que ya trae Compras.
COLUMNA_DESDE_COMPRAS: dict[str, str] = {"factem_edi": "fact_empaq"}

# Columnas de Compras que se calculan a partir de las anteriores.
COLUMNAS_CALCULADAS = ("ctobto_edi", "impart_edi")

# Columnas de control de la doble validacion (no forman parte del entregable).
COLUMNAS_CONTROL = ("impiva_edi_formula", "imieps_edi_formula")

_NO_ALFANUMERICO = re.compile(r"[^0-9A-Z]")
_NO_DIGITO = re.compile(r"\D")


@dataclass(slots=True)
class ResultadoCruce:
    """Metricas de una corrida del cruce."""

    df: pd.DataFrame
    renglones: int = 0
    cruzados_por_serie: int = 0
    cruzados_por_folio: int = 0
    celdas_llenadas: int = 0
    descartados_ambiguos: int = 0
    discrepancias_iva: int = 0
    discrepancias_ieps: int = 0

    @property
    def cruzados(self) -> int:
        """Renglones que encontraron su CFDI (con hueco o no)."""
        return self.cruzados_por_serie + self.cruzados_por_folio

    @property
    def tasa_cruce(self) -> float:
        return self.cruzados / self.renglones if self.renglones else 0.0

    def resumen(self) -> str:
        lineas = [
            f"Renglones de Compras         : {self.renglones:,}",
            f"Con CFDI encontrado          : {self.cruzados:,} ({self.tasa_cruce:.1%})",
            f"  - por serie + folio        : {self.cruzados_por_serie:,}",
            f"  - por folio (respaldo)     : {self.cruzados_por_folio:,}",
            f"Sin CFDI (se dejan como están): {self.renglones - self.cruzados:,}",
            f"Celdas vacías rellenadas     : {self.celdas_llenadas:,}",
            f"CFDI en conflicto (no usados): {self.descartados_ambiguos:,}",
            "Doble validacion (copiado vs. formula sobre totfactura):",
            f"  - difieren en IVA          : {self.discrepancias_iva:,}",
            f"  - difieren en IEPS         : {self.discrepancias_ieps:,}",
        ]
        return "\n".join(lineas)


def normalizar_factura(valor: object) -> str:
    """`FN-21226` y `FN21226` colapsan a `FN21226`; util para cruzar series."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return _NO_ALFANUMERICO.sub("", str(valor).upper())


def solo_digitos(valor: object) -> str:
    """Parte numerica de una factura, para el cruce de respaldo."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return _NO_DIGITO.sub("", str(valor)).lstrip("0")


def leer_compras(ruta: Path | str) -> pd.DataFrame:
    """Lee el Compras completo, uniendo todas sus hojas de datos.

    El Compras trae una hoja por año ("Compras 2020", "Compras 2021", ...); leer solo una
    perderia renglones. Recorre todas las hojas "Compras*", toma en cada una la fila de
    encabezados (la que empieza en `cnpj`) y concatena los datos que siguen. Excluye
    "Pendientes_EDI".
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        hojas = [n for n in wb.sheetnames if str(n).startswith("Compras")]
        if not hojas:
            hojas = [n for n in wb.sheetnames if n != "Pendientes_EDI"] or list(wb.sheetnames)

        columnas: list[str] | None = None
        datos: list[tuple] = []
        for nombre in hojas:
            filas = wb[nombre].iter_rows(values_only=True)
            encabezado = next(
                (list(fila) for fila in filas if fila and fila[0] == "cnpj"), None
            )
            if encabezado is None:
                continue  # hoja sin datos de Compras
            if columnas is None:
                columnas = [str(c) if c is not None else "" for c in encabezado]
            # `filas` ya quedo posicionado despues del encabezado de esta hoja.
            datos.extend(fila[: len(columnas)] for fila in filas)
        if columnas is None:
            raise ValueError(f"No se encontro la fila de encabezados en {ruta}")
    finally:
        wb.close()
    return pd.DataFrame(datos, columns=columnas)


def rfc_de_compras(compras: pd.DataFrame) -> str:
    """RFC del proveedor, leido de la columna `cnpj` del archivo de Compras.

    Evita que el usuario tenga que saber el RFC a partir del numero de proveedor:
    el Compras ya lo trae. Ver docs/PLANEACION.md seccion 5.
    """
    if "cnpj" not in compras.columns:
        return ""
    valores = compras["cnpj"].dropna().astype(str).str.strip()
    valores = valores[valores != ""]
    return valores.iloc[0] if not valores.empty else ""


# Deduplicación compartida: por cada (rfc, year) elige el `request_id` con mayor cobertura
# (más años distintos, y a igualdad más filas). Un mismo RFC puede tener varias descargas
# con años que se solapan —p. ej. una de 2020-2025 y otra de solo 2025—; esto evita
# contarlos dos veces. Consume los parámetros [patron, rfc, patron, rfc] en ese orden.
# Lo usa `cargar_cpa` para leer los datos ya deduplicados. El soporte (qué ZIP copiar)
# no usa este CTE sino `request_id_principal`, que elige una sola descarga.
_CTE_ELEGIDO = """
    WITH cobertura AS (
        SELECT request_id, COUNT(DISTINCT year) AS anios, COUNT(*) AS filas
        FROM read_parquet(?, hive_partitioning = 1)
        WHERE rfc = ?
        GROUP BY request_id
    ),
    elegido AS (
        SELECT year, request_id FROM (
            SELECT d.year, d.request_id,
                   ROW_NUMBER() OVER (
                       PARTITION BY d.year
                       ORDER BY c.anios DESC, c.filas DESC, d.request_id
                   ) AS rn
            FROM (SELECT DISTINCT year, request_id
                  FROM read_parquet(?, hive_partitioning = 1)
                  WHERE rfc = ?) d
            JOIN cobertura c USING (request_id)
        ) WHERE rn = 1
    )
"""


def cargar_cpa(
    rfc: str,
    parquet_root: Path | str,
    barcodes: set[str] | None = None,
    facturas: tuple[set[str], set[str]] | None = None,
) -> pd.DataFrame:
    """Lee del dataset Parquet los conceptos de un proveedor, ya deduplicados.

    La deduplicación por `request_id` está en `_CTE_ELEGIDO`.

    `barcodes` restringe la lectura a los códigos de barras del Compras (ya sin ceros a
    la izquierda). `facturas` (opcional) es `(series_folios, folios)` con las llaves de
    factura del Compras ya normalizadas —serie+folio y solo dígitos—: filtra la CPA a las
    facturas presentes en este lote. En proveedores con CPA enorme (Pepsico, 16M) unos
    pocos códigos de barra matchean millones de conceptos; el filtro por factura recorta eso
    a lo que de verdad puede cruzar y es lo que lo hace caber en memoria.
    """
    import duckdb

    columnas = ", ".join(f'"{origen}"' for origen in COLUMNAS_COPIADAS.values())
    patron = f"{Path(parquet_root).as_posix()}/**/*.parquet"
    # Codigo de barras normalizado igual que solo_digitos(): sin no-digitos ni ceros iniciales.
    barcode_sql = "ltrim(regexp_replace(p.\"noIdentificacion\", '[^0-9]', '', 'g'), '0')"
    # Factura normalizada igual que normalizar_factura()/solo_digitos() en Python.
    serie_folio_sql = (
        "regexp_replace(upper(coalesce(CAST(p.\"Serie\" AS VARCHAR),'') || "
        "coalesce(CAST(p.\"Folio\" AS VARCHAR),'')), '[^0-9A-Z]', '', 'g')"
    )
    folio_sql = "ltrim(regexp_replace(CAST(p.\"Folio\" AS VARCHAR), '[^0-9]', '', 'g'), '0')"

    consulta = _CTE_ELEGIDO + f"""
        SELECT p."Serie", p."Folio", p."noIdentificacion", {columnas}
        FROM read_parquet(?, hive_partitioning = 1) p
        JOIN elegido e ON p.year = e.year AND p.request_id = e.request_id
        WHERE p.rfc = ?
    """
    parametros = [patron, rfc, patron, rfc, patron, rfc]
    if barcodes:
        consulta += f" AND {barcode_sql} IN (SELECT bc FROM codigos)"
    if facturas:
        # Cualquiera de las dos llaves del cruce: serie+folio (principal) o solo dígitos
        # (respaldo). Con OR no se pierde ningún match posible.
        consulta += (
            f" AND ({serie_folio_sql} IN (SELECT sf FROM series_folios)"
            f" OR {folio_sql} IN (SELECT fo FROM folios_fac))"
        )

    with duckdb.connect() as con:
        _limitar_memoria_duckdb(con)
        if barcodes:
            con.register("codigos", pd.DataFrame({"bc": sorted(barcodes)}))
        if facturas:
            series_folios, folios = facturas
            con.register("series_folios", pd.DataFrame({"sf": sorted(series_folios - {""})}))
            con.register("folios_fac", pd.DataFrame({"fo": sorted(folios - {""})}))
        return con.execute(consulta, parametros).df()


def _limitar_memoria_duckdb(con) -> None:
    """Acota la RAM de DuckDB y le da carpeta de spill a disco.

    En proveedores grandes un año puede matchear millones de conceptos de CPA (Arca 2025:
    230 códigos de barra pero 2.9M filas de CPA). Sin límite, DuckDB intenta materializar
    todo en RAM y muere junto al año de compras que ya está en memoria. Con el límite,
    DuckDB usa disco para lo que no quepa y el cruce sobrevive.
    """
    import tempfile

    con.execute("PRAGMA disable_progress_bar")
    try:
        con.execute("SET memory_limit='4GB'")
        con.execute(f"SET temp_directory='{tempfile.gettempdir()}'")
    except Exception:  # noqa: BLE001 — si la versión de DuckDB no lo soporta, seguir igual
        pass


@dataclass(slots=True)
class SoporteRequest:
    """Qué descarga de CPA Vision respalda a un RFC, ya elegida una sola."""

    principal: str | None  # el request_id más completo: el ÚNICO que se copia como soporte
    anios_principal: list[str]  # años que cubre el principal
    anios_totales: list[str]  # años presentes para el RFC (unión de todas las descargas)
    redundantes: list[str]  # otras descargas del mismo RFC, que NO se copian

    @property
    def cobertura_completa(self) -> bool:
        """True si el principal cubre por sí solo todos los años del RFC."""
        return self.principal is not None and set(self.anios_principal) >= set(self.anios_totales)


def request_id_principal(rfc: str, parquet_root: Path | str) -> SoporteRequest:
    """Elige UNA sola descarga como soporte del RFC: la más completa.

    Si un proveedor se descargó dos veces (p. ej. una vez 2020-2025 y otra solo 2025), las
    dos descargas contienen la MISMA información de los años que se solapan. Para no guardar
    ese soporte duplicado se copia una sola: la de mayor cobertura (más años distintos y, a
    igualdad, más filas). Las demás se reportan como redundantes y no se copian.

    Devuelve también los años del principal frente a los del RFC, para poder avisar en el
    caso —que hoy no ocurre— en que ninguna descarga cubra sola todo el periodo.
    """
    import duckdb

    patron = f"{Path(parquet_root).as_posix()}/**/*.parquet"
    with duckdb.connect() as con:
        con.execute("PRAGMA disable_progress_bar")
        cobertura = con.execute(
            """
            SELECT request_id, COUNT(DISTINCT year) AS anios, COUNT(*) AS filas
            FROM read_parquet(?, hive_partitioning = 1)
            WHERE rfc = ?
            GROUP BY request_id
            ORDER BY anios DESC, filas DESC, request_id
            """,
            [patron, rfc],
        ).fetchall()
        if not cobertura:
            return SoporteRequest(None, [], [], [])

        principal = str(cobertura[0][0])
        redundantes = [str(fila[0]) for fila in cobertura[1:]]
        anios_principal = [
            str(fila[0])
            for fila in con.execute(
                "SELECT DISTINCT year FROM read_parquet(?, hive_partitioning = 1) "
                "WHERE rfc = ? AND request_id = ? ORDER BY 1",
                [patron, rfc, principal],
            ).fetchall()
        ]
        anios_totales = [
            str(fila[0])
            for fila in con.execute(
                "SELECT DISTINCT year FROM read_parquet(?, hive_partitioning = 1) "
                "WHERE rfc = ? ORDER BY 1",
                [patron, rfc],
            ).fetchall()
        ]
    return SoporteRequest(principal, anios_principal, anios_totales, redundantes)


def cruzar_proveedor(
    compras_path: Path | str, parquet_root: Path | str, *, rfc: str | None = None
) -> tuple[ResultadoCruce, str]:
    """Orquesta el cruce de un proveedor de punta a punta.

    Lee el Compras, resuelve el RFC (del argumento o de `cnpj`), carga de CPA Vision
    solo los conceptos cuyos códigos de barras aparecen en el Compras y cruza. Devuelve
    el resultado y el RFC usado. Es lo que llaman la GUI y el CLI.
    """
    compras = leer_compras(compras_path)
    rfc = (rfc or rfc_de_compras(compras)).upper()
    if not rfc:
        raise ValueError("No se pudo determinar el RFC (ni argumento ni columna cnpj).")
    barcodes = set(compras["codbarra"].map(solo_digitos)) - {""}
    cpa = cargar_cpa(rfc, parquet_root, barcodes=barcodes)
    return cruzar(compras, cpa), rfc


def cruzar(compras: pd.DataFrame, cpa: pd.DataFrame, *, en_sitio: bool = False) -> ResultadoCruce:
    """Llena las columnas EDI de `compras` con los conceptos de `cpa`.

    Con `en_sitio=True` **muta** `compras` y `cpa` en vez de copiarlos; quien llama debe
    poseerlos y no volver a usarlos. Lo usa el pipeline por año, donde cada año es un df
    propio: ahorrar la copia de `compras` es lo que hace caber un año grande en memoria.
    """
    df = compras if en_sitio else compras.copy()
    resultado = ResultadoCruce(df=df, renglones=len(df))

    df["_barcode"] = df["codbarra"].map(solo_digitos)
    df["_serie_folio"] = df["invnbr"].map(normalizar_factura)
    df["_folio"] = df["invnbr"].map(solo_digitos)

    if not en_sitio:
        cpa = cpa.copy()
    cpa["_barcode"] = cpa["noIdentificacion"].map(solo_digitos)
    cpa["_serie_folio"] = (
        cpa["Serie"].fillna("").astype(str) + cpa["Folio"].fillna("").astype(str)
    ).map(normalizar_factura)
    cpa["_folio"] = cpa["Folio"].map(solo_digitos)

    por_serie, ambiguos_serie = _indexar(cpa, ["_barcode", "_serie_folio"])
    por_folio, ambiguos_folio = _indexar(cpa, ["_barcode", "_folio"])
    resultado.descartados_ambiguos = ambiguos_serie + ambiguos_folio

    origen = _buscar(df, por_serie, ["_barcode", "_serie_folio"])
    falta = origen[list(COLUMNAS_COPIADAS)].isna().all(axis=1)
    resultado.cruzados_por_serie = int((~falta).sum())

    if falta.any():
        respaldo = _buscar(df.loc[falta], por_folio, ["_barcode", "_folio"])
        origen.loc[falta] = respaldo
        cruzo_respaldo = ~respaldo[list(COLUMNAS_COPIADAS)].isna().all(axis=1)
        resultado.cruzados_por_folio = int(cruzo_respaldo.sum())

    # El cruce SOLO rellena celdas vacias: nunca pisa un dato que Compras ya traia.
    cruzo = origen[list(COLUMNAS_COPIADAS)].notna().any(axis=1)

    for destino in COLUMNAS_COPIADAS:
        df[destino] = df[destino].astype(object) if destino in df.columns else pd.NA
        hueco = _vacio(df[destino]) & origen[destino].notna()
        df.loc[hueco, destino] = origen.loc[hueco, destino]
        resultado.celdas_llenadas += int(hueco.sum())

    for destino, fuente in COLUMNA_DESDE_COMPRAS.items():
        df[destino] = df[destino].astype(object) if destino in df.columns else pd.NA
        hueco = _vacio(df[destino]) & cruzo
        df.loc[hueco, destino] = df.loc[hueco, fuente]
        resultado.celdas_llenadas += int(hueco.sum())

    _calcular_derivadas(df, resultado)
    _validar_importes(df, resultado)

    df.drop(columns=[c for c in df.columns if c.startswith("_")], inplace=True)
    return resultado


def _vacio(serie: pd.Series) -> pd.Series:
    """True donde la celda está sin dato (nulo o cadena en blanco)."""
    return serie.isna() | (serie.astype(str).str.strip().isin(["", "None", "nan"]))


def _indexar(cpa: pd.DataFrame, llave: list[str]) -> tuple[pd.DataFrame, int]:
    """Deja un solo concepto por llave y reporta cuantos se descartaron por conflicto.

    Un mismo producto suele venir en varias líneas de la misma factura con **valores
    idénticos** (medido en Nestlé: 99.94% de las llaves repetidas coinciden en todo lo
    que se copia). Esas se colapsan en una sola: dan igual. Solo se descarta la llave
    cuando los conceptos **difieren** en algún valor copiado — ahí no se puede elegir
    sin una regla de negocio, y preferimos no llenar a llenar mal.
    """
    valido = (cpa[llave] != "").all(axis=1)
    datos = cpa.loc[valido]
    # Colapsa las lineas identicas en llave + todo lo que se copia (duplicados inofensivos).
    datos = datos.drop_duplicates(subset=llave + list(COLUMNAS_COPIADAS.values()))
    # Lo que siga repetido en la llave es un conflicto real de valores: se descarta.
    unico = datos.drop_duplicates(subset=llave, keep=False)
    descartados = int(len(datos) - len(unico))
    return unico.set_index(llave), descartados


def _buscar(df: pd.DataFrame, indice: pd.DataFrame, llave: list[str]) -> pd.DataFrame:
    """Trae los valores del CFDI alineados con el indice de `df`."""
    encontrado = indice.reindex(pd.MultiIndex.from_frame(df[llave]))
    encontrado.index = df.index
    return encontrado.rename(
        columns={origen: destino for destino, origen in COLUMNAS_COPIADAS.items()}
    )


def _calcular_derivadas(df: pd.DataFrame, resultado: ResultadoCruce) -> None:
    """`ctobto_edi` y `impart_edi` según las fórmulas del archivo de Luis.

    Solo se escriben donde estaban vacías y donde ya hay insumo (`ctonto_edi`), para no
    pisar lo que Compras traía ni inventar ceros en renglones sin CFDI.
    """
    ctonto = to_number(df["ctonto_edi"])
    factem = to_number(df["factem_edi"])
    canfac = to_number(df["canfac_edi"])
    poriva = to_number(df["poriva_edi"])
    con_insumo = _vacio(df["ctonto_edi"]).__invert__()

    ctobto = ctonto * factem
    impart = ctobto * canfac * (1 + poriva)
    for destino, valores in (("ctobto_edi", ctobto), ("impart_edi", impart)):
        df[destino] = df[destino].astype(object) if destino in df.columns else pd.NA
        hueco = _vacio(df[destino]) & con_insumo
        df.loc[hueco, destino] = valores[hueco]
        resultado.celdas_llenadas += int(hueco.sum())


def _validar_importes(df: pd.DataFrame, resultado: ResultadoCruce) -> None:
    """Doble validacion: importe copiado del CFDI contra la formula sobre totfactura.

    La formula se conserva solo como control. El valor bueno es el del CFDI:
    ver `docs/MAPEO_CRUCE_CPA_COMPRAS.md` seccion 4.0.2.1.
    """
    total = to_number(df["totfactura"])
    for destino, tasa_col in (("impiva_edi", "poriva_edi"), ("imieps_edi", "prieps_edi")):
        tasa = to_number(df[tasa_col])
        formula = (total / (1 + tasa) * tasa).where(tasa > 0, 0.0)
        df[f"{destino}_formula"] = formula
        difieren = (to_number(df[destino]) - formula).abs() > 0.02
        if destino == "impiva_edi":
            resultado.discrepancias_iva = int(difieren.sum())
        else:
            resultado.discrepancias_ieps = int(difieren.sum())
