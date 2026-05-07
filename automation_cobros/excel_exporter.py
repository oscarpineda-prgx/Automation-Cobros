from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
import config
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from automation_cobros.calculations import (
    COMPRAS_COLUMNS,
    EDI_COLUMNS,
    build_pending_edi_dataframe,
    prepare_compras_dataframe,
)
from automation_cobros.utils import ensure_parent


HEADER_ROW = 7
DATA_ROW = 8

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="FF00FD28")
FORMULA_FILL = PatternFill("solid", fgColor="FFFFFF00")
EDI_FILL = PatternFill("solid", fgColor="FFF2CC")
AUDIT_FILL = PatternFill("solid", fgColor="E2F0D9")
PENDING_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="FF000000", bold=True)
TITLE_FONT = Font(color="FF000000", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

FORMULA_COLUMNS = {
    "concaten",
    "fante",
    "facdecto",
    "ctouni_sistema",
    "ctontol",
    "impaud",
    "dpagar",
    "imp",
    "dif cto fac ctouni",
    "ctontopza",
}


def write_compras_workbook(
    df: pd.DataFrame,
    output_path: Path,
    vendor: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> Path:
    output_path = Path(output_path)
    ensure_parent(output_path)

    prepared = prepare_compras_dataframe(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    _add_logo(ws)
    _write_title(ws, prepared, vendor, start_date, end_date)
    _write_dataframe(ws, prepared, COMPRAS_COLUMNS, HEADER_ROW)
    _style_compras_sheet(ws, prepared)
    _apply_formula_columns(ws, len(prepared))
    _write_impaud_helper_sheet(wb, prepared)
    _write_pending_sheet(wb, prepared)
    _force_formula_recalculation(wb)

    wb.save(output_path)
    return output_path


def _write_title(ws, df: pd.DataFrame, vendor: str | None, start_date: str | None, end_date: str | None) -> None:
    vendor_name = ""
    if "vndname" in df.columns and not df.empty:
        vendor_name = str(df["vndname"].dropna().iloc[0]) if df["vndname"].notna().any() else ""
    vendor_code = vendor or (str(df["vndnbr"].dropna().iloc[0]) if "vndnbr" in df.columns and df["vndnbr"].notna().any() else "")
    period = _period_label(start_date, end_date)

    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=12)
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=12)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=12)
    ws["D2"] = "Tiendas Soriana, S.A. de C.V."
    ws["D3"] = f"{vendor_code} - {vendor_name}".strip(" -")
    ws["D4"] = f"Compras Periodo {period}"
    for cell in ["D2", "D3", "D4"]:
        ws[cell].font = TITLE_FONT
        ws[cell].alignment = Alignment(horizontal="center")
    for row_idx in range(2, 5):
        ws.row_dimensions[row_idx].height = 21


def _add_logo(ws) -> None:
    logo_path = config.BASE_DIR / "templates" / "Soriana-Logo.png"
    if not logo_path.exists():
        return
    logo = XLImage(str(logo_path))
    logo.width = 190
    logo.height = 46
    ws.add_image(logo, "A2")


def _write_dataframe(ws, df: pd.DataFrame, columns: Iterable[str], header_row: int) -> None:
    ws.row_dimensions[header_row].height = 20
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(header_row, col_idx, column)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, _excel_value(value))
            cell.border = THIN_BORDER


def _style_compras_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A8"
    last_col = get_column_letter(len(COMPRAS_COLUMNS))
    last_row = HEADER_ROW + max(len(df), 1)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_row}"
    ws.sheet_view.showGridLines = False

    column_index = {column: idx for idx, column in enumerate(COMPRAS_COLUMNS, start=1)}
    for edi_col in EDI_COLUMNS:
        if edi_col in column_index:
            ws.cell(HEADER_ROW, column_index[edi_col]).fill = HEADER_FILL
            ws.cell(HEADER_ROW, column_index[edi_col]).font = HEADER_FONT

    for audit_col in [
        "cto_aud",
        "iva_aud",
        "ieps_aud",
        "imp_aud",
        "debio_pagar_ne",
        "dif_det_ne",
        "debio_pagar_inv",
        "tot_pagado_inv",
        "dif_det_inv",
    ]:
        if audit_col in column_index:
            ws.cell(HEADER_ROW, column_index[audit_col]).fill = HEADER_FILL
            ws.cell(HEADER_ROW, column_index[audit_col]).font = HEADER_FONT

    widths = {
        "cnpj": 16,
        "vndnbr": 10,
        "vndname": 28,
        "ponbr": 14,
        "podt": 12,
        "rcvnbr": 12,
        "rcvdt": 12,
        "strnbr": 12,
        "invnbr": 18,
        "itmdesc": 34,
        "uuid": 36,
        "txt_cabec": 24,
        "txt_item": 24,
    }
    for idx, column in enumerate(COMPRAS_COLUMNS, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(column, 13)
        if column in _date_columns():
            _set_number_format(ws, idx, "yyyy-mm-dd")
        elif column in _money_columns():
            _set_number_format(ws, idx, '#,##0.00')
        elif column in _percent_columns():
            _set_number_format(ws, idx, "0.00%")

    for row in ws.iter_rows(min_row=DATA_ROW, max_row=DATA_ROW + len(df) - 1):
        for cell in row:
            cell.alignment = Alignment(vertical="center")


def _apply_formula_columns(ws, row_count: int) -> None:
    column_index = {column: idx for idx, column in enumerate(COMPRAS_COLUMNS, start=1)}
    for column in FORMULA_COLUMNS:
        col_idx = column_index[column]
        header = ws.cell(HEADER_ROW, col_idx)
        header.fill = FORMULA_FILL
        header.font = HEADER_FONT

    if row_count <= 0:
        return

    letters = {column: get_column_letter(idx) for column, idx in column_index.items()}
    for row_idx in range(DATA_ROW, DATA_ROW + row_count):
        formulas = {
            "concaten": f"=CONCATENATE({letters['strnbr']}{row_idx},{letters['rcvnbr']}{row_idx})",
            "fante": f"=({letters['canfac_edi']}{row_idx}*{letters['fact_empaq']}{row_idx})-{letters['can_rec']}{row_idx}",
            "facdecto": f"=100-({letters['poitmnetcst']}{row_idx}/{letters['poitmgrscst']}{row_idx})*100",
            "ctouni_sistema": f"=({letters['poitmgrscst']}{row_idx}/{letters['fact_empaq']}{row_idx})*(1-{letters['facdecto']}{row_idx}/100)",
            "ctontol": f"=IF(AND({letters['ctonto_edi']}{row_idx}<{letters['ctouni_sistema']}{row_idx},{letters['ctonto_edi']}{row_idx}>0),{letters['ctonto_edi']}{row_idx},{letters['ctouni_sistema']}{row_idx})",
            "impaud": f"=({letters['ctontol']}{row_idx}*{letters['can_rec']}{row_idx})*(1+{letters['imp']}{row_idx})",
            "dpagar": f"=VLOOKUP({letters['concaten']}{row_idx},impaud!$A$3:$B$4779,2,0)",
            "imp": "=0.16",
            "dif cto fac ctouni": f"={letters['ctonto_edi']}{row_idx}-{letters['ctontol']}{row_idx}",
            "ctontopza": f"={letters['ctonto_edi']}{row_idx}/{letters['fact_empaq']}{row_idx}",
        }
        for column, formula in formulas.items():
            cell = ws.cell(row_idx, column_index[column])
            cell.value = formula


def _force_formula_recalculation(wb: Workbook) -> None:
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


def _write_impaud_helper_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("impaud")
    ws.sheet_state = "hidden"
    ws["A2"] = "concaten"
    ws["B2"] = "impaud"

    if "concaten" not in df.columns or df.empty:
        return

    unique_keys = df["concaten"].dropna().astype(str).drop_duplicates().tolist()
    max_lookup_rows = 4779 - 3 + 1
    for offset, key in enumerate(unique_keys[:max_lookup_rows], start=3):
        ws.cell(offset, 1, key)
        ws.cell(offset, 2, f"=SUMIF(Compras!$D:$D,A{offset},Compras!$AO:$AO)")


def _write_pending_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    pending = build_pending_edi_dataframe(df)
    ws = wb.create_sheet("Pendientes_EDI")
    ws["A1"] = "Registros con campos EDI vacios para complemento manual o futura automatizacion PMAIL/CPA Vision"
    ws["A1"].font = Font(bold=True, color="9C5700")
    if pending.empty:
        ws["A3"] = "No se detectaron pendientes EDI."
        ws.column_dimensions["A"].width = 80
        return
    _write_dataframe(ws, pending, list(pending.columns), 3)
    for col_idx, column in enumerate(pending.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(36, len(str(column)) + 4))
        if column in _date_columns():
            _set_number_format(ws, col_idx, "m/d/yyyy", min_row=4)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A3:{get_column_letter(len(pending.columns))}{3 + len(pending)}"


def _excel_value(value):
    if pd.isna(value):
        return None
    return value


def _set_number_format(ws, col_idx: int, number_format: str, min_row: int = DATA_ROW) -> None:
    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=min_row):
        for item in cell:
            item.number_format = number_format


def _period_label(start_date: str | None, end_date: str | None) -> str:
    if not start_date and not end_date:
        return ""
    start = str(start_date or "")[:4]
    end = str(end_date or "")[:4]
    parsed_end = pd.to_datetime(end_date, errors="coerce") if end_date else None
    parsed_start = pd.to_datetime(start_date, errors="coerce") if start_date else None
    if parsed_end is not None and not pd.isna(parsed_end) and parsed_end.month == 1 and parsed_end.day == 1:
        if parsed_start is None or pd.isna(parsed_start) or parsed_end.year > parsed_start.year:
            end = str(parsed_end.year - 1)
    if start and end and start != end:
        return f"{start}-{end}"
    return start or end


def _date_columns() -> set[str]:
    return {"podt", "rcvdt", "invdt", "paychkdt", "payinvdt", "invdt_ne", "paychkdt_ne"}


def _money_columns() -> set[str]:
    return {
        "ctouni",
        "ctouni_sistema",
        "ctonto_edi",
        "ctontopza",
        "impart_edi",
        "imieps_edi",
        "impiva_edi",
        "totfactura",
        "cto_aud",
        "imp_aud",
        "debio_pagar_ne",
        "dif_det_ne",
        "debio_pagar_inv",
        "tot_pagado_inv",
        "dif_det_inv",
        "paynetamt",
        "tot_pagado_ne",
        "compra_bruta",
        "compra_bruta mas impuestos",
        "compra_neta",
        "compra neta mas impuestos",
    }


def _percent_columns() -> set[str]:
    return {"prieps_edi", "poriva_edi", "ieps_t007s", "iva_t007s", "iva_aud", "ieps_aud"}
