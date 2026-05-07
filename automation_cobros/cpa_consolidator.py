from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from zipfile import ZipFile

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from automation_cobros.utils import ensure_parent


EXCEL_MAX_ROWS = 1_048_576
DATA_SHEET = "Consolidado"
SUMMARY_SHEET = "Resumen"
HEADER_FILL = PatternFill("solid", fgColor="FF00FD28")
HEADER_FONT = Font(color="FF000000", bold=True)


@dataclass(slots=True)
class ConsolidationResult:
    output_path: Path
    rows: int
    source_files: int
    zip_files: int


def consolidate_cpa_zip_to_excel(input_path: Path | str, output_path: Path | str) -> ConsolidationResult:
    zip_paths = _resolve_zip_paths(Path(input_path))
    if not zip_paths:
        raise FileNotFoundError(f"No se encontraron archivos .zip en: {input_path}")

    frames: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []
    for zip_path in zip_paths:
        zip_frames, zip_summary = _read_zip_csvs(zip_path)
        frames.extend(zip_frames)
        summary_rows.extend(zip_summary)

    if not frames:
        raise ValueError("No se encontraron CSV dentro de los ZIP seleccionados.")

    consolidated = pd.concat(frames, ignore_index=True, sort=False)
    if len(consolidated) > EXCEL_MAX_ROWS - 1:
        raise ValueError(
            f"El consolidado tiene {len(consolidated):,} filas y excede el limite de Excel "
            f"({EXCEL_MAX_ROWS - 1:,}). Para ese volumen conviene SQLite."
        )

    summary = pd.DataFrame(summary_rows)
    output_path = Path(output_path)
    ensure_parent(output_path)
    _write_excel(consolidated, summary, output_path)
    return ConsolidationResult(
        output_path=output_path,
        rows=len(consolidated),
        source_files=len(summary_rows),
        zip_files=len(zip_paths),
    )


def _resolve_zip_paths(input_path: Path) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".zip":
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.glob("*.zip"), key=lambda path: path.stat().st_mtime, reverse=True)
    return []


def _read_zip_csvs(zip_path: Path) -> tuple[list[pd.DataFrame], list[dict[str, object]]]:
    frames: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []
    with ZipFile(zip_path) as zf:
        csv_names = [name for name in zf.namelist() if name.lower().endswith(".csv")]
        for csv_name in csv_names:
            with zf.open(csv_name) as csv_file:
                df = _read_csv(csv_file)
            df.columns = [_clean_column_name(column) for column in df.columns]
            year = _extract_year(csv_name)
            df.insert(0, "archivo_zip", zip_path.name)
            df.insert(1, "archivo_csv", Path(csv_name).name)
            df.insert(2, "anio", year)
            frames.append(df)
            summary_rows.append(
                {
                    "archivo_zip": zip_path.name,
                    "archivo_csv": Path(csv_name).name,
                    "anio": year,
                    "filas": len(df),
                    "columnas": len(df.columns),
                }
            )
    return frames, summary_rows


def _read_csv(csv_file) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "latin1", "cp1252"):
        try:
            return pd.read_csv(
                csv_file,
                sep=",",
                encoding=encoding,
                dtype=str,
                keep_default_na=False,
                low_memory=False,
            )
        except UnicodeDecodeError:
            csv_file.seek(0)
    csv_file.seek(0)
    return pd.read_csv(csv_file, sep=",", dtype=str, keep_default_na=False, low_memory=False)


def _clean_column_name(column: object) -> str:
    return str(column).replace("\ufeff", "").strip()


def _extract_year(filename: str) -> str:
    match = re.search(r"(20\d{2})", filename)
    return match.group(1) if match else ""


def _write_excel(consolidated: pd.DataFrame, summary: pd.DataFrame, output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
        consolidated.to_excel(writer, sheet_name=DATA_SHEET, index=False)

        _style_sheet(writer.book[SUMMARY_SHEET], summary)
        _style_sheet(writer.book[DATA_SHEET], consolidated)


def _style_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = _suggest_widths(df)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _suggest_widths(df: pd.DataFrame) -> list[int]:
    widths: list[int] = []
    sample = df.head(500)
    for column in df.columns:
        values = [len(str(column))]
        if not sample.empty:
            values.extend(sample[column].astype(str).str.len().tolist())
        widths.append(max(10, min(45, max(values) + 2)))
    return widths
