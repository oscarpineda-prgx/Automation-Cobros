"""Validaciones de Condiciones de Arca (391250) por AÑO, acotadas a una lista de folios.

CASO AISLADO (pedido de Oscar, 2026-08-11). No es parte del pipeline normal: Arca ya tenía
su Validación consolidada y aquí se pide otra cosa, partida por año y recortada a los
folios de "VF mayor a $20MX Dis Arca.xlsx".

Cómo se arma cada archivo
-------------------------
- **Consolidado**: sale TAL CUAL del archivo de folios. Ese archivo ya ES un Consolidado
  (trae las 21 columnas del formato más 3 de trabajo), así que no se recalcula nada: se
  reparte por año y se escribe. Recalcularlo solo introduciría diferencias contra lo que
  Oscar ya revisó.
- **Resumen**: suma de `Diferencia Ajustada` de los folios de ese año (decisión de Oscar:
  la cifra neta de devoluciones, $119.38M en total, no la bruta de $122.35M).
- **Detalle PAGOS**: los renglones de artículo de esos folios, leídos de los Compras
  trimestrales que generaron la Validación original, más una columna **Trimestre**.

Por qué el año y el trimestre salen de `rcvdt` (Fec Recibo)
-----------------------------------------------------------
Los Compras trimestrales se generaron llamando a `F_COMPRAS(proveedor, ini, fin)` por
trimestre. Verificado leyendo el 2020-T1: las fechas de recibo caen exactas dentro del
trimestre (2020-01-02..2020-03-31) mientras que las de pedido se salen (2019-12-29). O sea
la función filtra por fecha de recibo, y por lo tanto es `rcvdt` quien decide en qué archivo
—y en qué trimestre— vive cada renglón. La columna `Trimestre` se toma del archivo de origen,
que es la respuesta literal a "a qué trimestre corresponde con respecto a la base de compras".

Uso
---
    python scripts/arca_validacion_por_anio.py --extraer    # lee los Compras (lento, 1 vez)
    python scripts/arca_validacion_por_anio.py --generar    # escribe los Excel (rápido)
    python scripts/arca_validacion_por_anio.py              # las dos

La extracción deja un caché en Parquet para poder repetir `--generar` sin volver a leer
los ~7 GB de Compras.
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

# La consola de Windows es cp1252 y revienta con acentos o flechas. Un print no puede
# tumbar una lectura de horas: se fuerza UTF-8 y lo que no entre se sustituye.
for _flujo in (sys.stdout, sys.stderr):
    try:
        _flujo.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

import pandas as pd
import xlsxwriter

import config
from automation_costos.utils import clean_code_series, ensure_parent
from automation_costos.validation_exporter import (
    DETALLE_COLUMNS,
    _DetalleStream,
    _encabezar_xlsx,
    _formatos_xlsx,
    _titulo_xlsx,
    _volcar_xlsx,
    build_detalle_rapido,
)

PROVEEDOR = "391250"
NOMBRE = "DISTRIBUIDORA ARCA CONTINENTAL S DE RL DE CV"
ETIQUETA = f"{PROVEEDOR} - {NOMBRE}"

COMPRAS_DIR = Path(
    r"X:\Soriana\00 - AUDITORIA 2020 - 2024\Proceso Validación de condiciones (Oscar Pineda)"
    rf"\{PROVEEDOR}_{NOMBRE}"
)
SALIDA_DIR = COMPRAS_DIR
FILTRO = RAIZ / "VF mayor a $20MX Dis Arca.xlsx"
# Un Parquet por trimestre en vez de uno solo: leer los 24 Compras son horas, y asi un
# fallo a la mitad no obliga a empezar de cero (se salta lo que ya esta escrito).
CACHE_DIR = RAIZ / "outputs" / "arca_detalle_filtrado"

# El Detalle se arma con `build_detalle_rapido`, que espera los nombres de Compras.
COLUMNAS_FUENTE = [
    "ponbr", "podt", "rcvnbr", "rcvdt", "invnbr", "invnbr_ne", "strnbr",
    "nombre_division", "po_groupdescrip", "grupoarticulo", "cltstyle", "codbarra",
    "itmdesc", "fact_empaq", "can_rec", "ctouni", "cto_aud",
    "prieps_edi", "ieps_aud", "poriva_edi", "iva_aud", "imp_aud",
]

DETALLE_COLUMNS_ARCA = [*DETALLE_COLUMNS, "Trimestre"]
_RE_TRIMESTRE = re.compile(r"_(\d{4}-T\d)\.xlsx$", re.IGNORECASE)

# El Compras trae columnas con tipos MEZCLADOS: `prieps_edi`, por ejemplo, tiene el numero 0
# en unas celdas y el texto "0" en otras (unas las escribio el exportador y otras el auditor
# a mano en Excel). Parquet exige un tipo por columna, asi que hay que decidirlo aqui; sin
# esto pyarrow aborta con "Could not convert '0' with type str: tried to convert to double".
COLUMNAS_NUMERICAS = (
    "fact_empaq", "can_rec", "ctouni", "cto_aud",
    "prieps_edi", "ieps_aud", "poriva_edi", "iva_aud", "imp_aud",
)
COLUMNAS_FECHA = ("podt", "rcvdt")


def normalizar_tipos(df: pd.DataFrame) -> pd.DataFrame:
    """Un tipo estable por columna. Idempotente: se aplica al escribir y al leer.

    Los mismos criterios que usa `build_detalle_rapido` mas adelante (`to_number` en los
    importes, texto en los codigos), solo que adelantados para que Parquet pueda guardarlos.
    """
    for columna in COLUMNAS_NUMERICAS:
        if columna in df.columns:
            df[columna] = pd.to_numeric(df[columna], errors="coerce")
    for columna in COLUMNAS_FECHA:
        if columna in df.columns:
            df[columna] = pd.to_datetime(df[columna], errors="coerce")
    resto = [c for c in df.columns if c not in COLUMNAS_NUMERICAS and c not in COLUMNAS_FECHA]
    for columna in resto:
        df[columna] = df[columna].astype("string")
    return df


# --------------------------------------------------------------------------------------
# Filtro
# --------------------------------------------------------------------------------------

def cargar_filtro() -> pd.DataFrame:
    """El Consolidado que pidió Oscar, con el año ya resuelto por Fec Recibo."""
    df = pd.read_excel(FILTRO, header=1)
    df = df[df["Folio"].notna()].copy()
    df["Folio"] = clean_code_series(df["Folio"])
    df["Fec Recibo"] = pd.to_datetime(df["Fec Recibo"], errors="coerce")
    df["_anio"] = df["Fec Recibo"].dt.year.astype("Int64")
    return df


# --------------------------------------------------------------------------------------
# Fase 1: extraer del Compras trimestral solo los renglones de esos folios
# --------------------------------------------------------------------------------------

def extraer(folios: set[str]) -> None:
    import openpyxl

    archivos = sorted(COMPRAS_DIR.glob(f"Compras_{PROVEEDOR}_*_20??-T?.xlsx"))
    if not archivos:
        raise SystemExit(f"No se encontraron Compras trimestrales en {COMPRAS_DIR}")

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"{len(archivos)} archivos de Compras | {len(folios):,} folios buscados\n", flush=True)
    t_total = time.time()

    for n, ruta in enumerate(archivos, start=1):
        etiqueta = _RE_TRIMESTRE.search(ruta.name).group(1)
        destino = CACHE_DIR / f"{etiqueta}.parquet"
        if destino.exists():
            print(f"  [{n:>2}/{len(archivos)}] {etiqueta}  ya extraido, se omite", flush=True)
            continue
        t0 = time.time()
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            hojas = [h for h in wb.sheetnames if h.startswith("Compras")] or [wb.sheetnames[0]]
            filas: list[tuple] = []
            leidas = 0
            for hoja in hojas:
                it = wb[hoja].iter_rows(values_only=True)
                encabezado = next((list(f) for f in it if f and f[0] == "cnpj"), None)
                if encabezado is None:
                    continue
                idx = {c: i for i, c in enumerate(encabezado)}
                faltan = [c for c in COLUMNAS_FUENTE if c not in idx]
                if faltan:
                    raise SystemExit(f"{ruta.name}: faltan columnas {faltan}")
                i_str, i_rcv = idx["strnbr"], idx["rcvnbr"]
                posiciones = [idx[c] for c in COLUMNAS_FUENTE]
                for fila in it:
                    leidas += 1
                    folio = _folio(fila[i_str], fila[i_rcv])
                    if folio in folios:
                        filas.append(tuple(fila[p] for p in posiciones))
        finally:
            wb.close()

        parte = pd.DataFrame(filas, columns=COLUMNAS_FUENTE)
        parte["Trimestre"] = etiqueta
        parte = normalizar_tipos(parte)
        parte.to_parquet(destino, index=False)
        print(
            f"  [{n:>2}/{len(archivos)}] {etiqueta}  |  {leidas:>9,} leidas  ->  "
            f"{len(parte):>8,} del filtro  |  {time.time() - t0:6.0f}s",
            flush=True,
        )

    total = sum(len(pd.read_parquet(p)) for p in sorted(CACHE_DIR.glob("*.parquet")))
    print(
        f"\nCache: {CACHE_DIR}  ({total:,} renglones, {time.time() - t_total:.0f}s)",
        flush=True,
    )


def _folio(tienda, nota) -> str:
    """Misma llave que `utils.make_folio`, sin pasar por pandas (se llama por renglón)."""
    t = "" if tienda is None else str(tienda).strip()
    r = "" if nota is None else str(nota).strip()
    if t.endswith(".0"):
        t = t[:-2]
    if r.endswith(".0"):
        r = r[:-2]
    return f"{config.FOLIO_PREFIX}{t.zfill(4)[-4:]}{r.zfill(8)[-8:]}"


# --------------------------------------------------------------------------------------
# Fase 2: escribir un Excel por año, con el estilo de siempre
# --------------------------------------------------------------------------------------

def generar(filtro: pd.DataFrame) -> None:
    partes = sorted(CACHE_DIR.glob("*.parquet"))
    if not partes:
        raise SystemExit(f"Falta el cache {CACHE_DIR}. Corre primero con --extraer.")
    # Se re-normaliza al leer: los primeros trimestres se escribieron antes de que existiera
    # `normalizar_tipos` y concatenarlos con los nuevos dejaria columnas de tipo mixto.
    detalle = normalizar_tipos(
        pd.concat([pd.read_parquet(p) for p in partes], ignore_index=True)
    )
    detalle["_folio"] = _folio_series(detalle)
    detalle = _quitar_colisiones(detalle, filtro)

    anios = sorted(a for a in filtro["_anio"].dropna().unique())
    print(f"\nAños a generar: {anios}\n", flush=True)
    resumen_global = []

    for anio in anios:
        cons = filtro[filtro["_anio"] == anio].drop(columns=["_anio"]).reset_index(drop=True)
        folios = set(cons["Folio"])
        det_anio = detalle[detalle["_folio"].isin(folios)]
        salida = SALIDA_DIR / f"Validacion_{PROVEEDOR}_{NOMBRE}_{anio}.xlsx"
        total = float(pd.to_numeric(cons["Diferencia Ajustada"], errors="coerce").fillna(0).sum())

        _escribir(salida, cons, det_anio, total)
        faltan = len(folios) - det_anio["_folio"].nunique()
        resumen_global.append(
            {
                "Anio": int(anio), "Folios": len(cons), "Renglones detalle": len(det_anio),
                "Folios sin detalle": faltan, "Diferencia Ajustada": total,
                "Trimestres": ", ".join(sorted(det_anio["Trimestre"].unique())),
            }
        )
        print(
            f"  {anio}  |  {len(cons):>6,} folios  |  {len(det_anio):>9,} renglones  |  "
            f"${total:>16,.2f}  |  sin detalle: {faltan}  ->  {salida.name}",
            flush=True,
        )

    tabla = pd.DataFrame(resumen_global)
    print("\n" + tabla.to_string(index=False))
    print(f"\nTOTAL Diferencia Ajustada: ${tabla['Diferencia Ajustada'].sum():,.2f}")
    print(f"TOTAL folios            : {tabla['Folios'].sum():,}")


def _quitar_colisiones(detalle: pd.DataFrame, filtro: pd.DataFrame) -> pd.DataFrame:
    """Descarta los renglones que cruzaron por folio pero son de OTRA nota de entrada.

    El folio es `prefijo + tienda + nota`, sin fecha, y el numero de nota se reutiliza:
    medido en Arca, `11004035400243772` existe como nota recibida el 2023-09-28 (la del
    filtro) y otra vez el 2024-03-27. Cruzar solo por folio arrastraria los renglones de la
    segunda al archivo de 2023.

    La desambiguacion es la fecha de recibo: 2,027,860 de 2,027,905 renglones (99.998%)
    traen `rcvdt` identico al `Fec Recibo` del filtro, y las 45 excepciones son exactamente
    los 4 folios repetidos. Asi que la llave real es folio + fecha de recibo.
    """
    fechas = dict(zip(filtro["Folio"], filtro["Fec Recibo"]))
    esperada = detalle["_folio"].map(fechas)
    sobra = detalle["rcvdt"].ne(esperada)
    if sobra.any():
        print(
            f"  Colisiones de folio descartadas: {int(sobra.sum())} renglones "
            f"de {detalle.loc[sobra, '_folio'].nunique()} folios reutilizados",
            flush=True,
        )
    return detalle[~sobra]


def _folio_series(detalle: pd.DataFrame) -> pd.Series:
    tienda = clean_code_series(detalle["strnbr"]).str.zfill(4).str[-4:]
    nota = clean_code_series(detalle["rcvnbr"]).str.zfill(8).str[-8:]
    return config.FOLIO_PREFIX + tienda + nota


def _escribir(salida: Path, consolidado: pd.DataFrame, detalle: pd.DataFrame, total: float) -> None:
    """Mismo estilo que la Validación de siempre: reusa los helpers del exportador."""
    ensure_parent(salida)
    columnas_cons = list(consolidado.columns)
    wb = xlsxwriter.Workbook(str(salida), {"constant_memory": True, "use_zip64": True})
    try:
        fmts = _formatos_xlsx(wb)

        resumen = wb.add_worksheet("Resumen")
        _titulo_xlsx(resumen, ETIQUETA, fmts)
        resumen.set_column(2, 2, 22)
        resumen.set_column(3, 3, 28)
        resumen.write(6, 2, "Resultado Auditoria", fmts["hdr"])
        resumen.write(6, 3, "Observaciones Auditor", fmts["hdr"])
        resumen.write_number(7, 2, total, fmts["money"])
        resumen.write(7, 3, "Diferencia costos" if len(consolidado) else "")

        _volcar_xlsx(wb, "Consolidado", columnas_cons, consolidado, ETIQUETA, fmts, con_totales=True)

        stream = _DetalleStream(wb, DETALLE_COLUMNS_ARCA, ETIQUETA, fmts)
        if not detalle.empty:
            det = build_detalle_rapido(detalle, [])
            det["Trimestre"] = detalle["Trimestre"].values
            stream.escribir(det[DETALLE_COLUMNS_ARCA])
        wb.close()
    except BaseException:
        try:
            wb.close()
        except BaseException:
            pass
        salida.unlink(missing_ok=True)
        raise


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--extraer", action="store_true", help="Solo leer los Compras y armar el caché")
    p.add_argument("--generar", action="store_true", help="Solo escribir los Excel desde el caché")
    args = p.parse_args()
    hacer_todo = not (args.extraer or args.generar)

    filtro = cargar_filtro()
    print(f"Filtro: {len(filtro):,} folios | anios {sorted(int(a) for a in filtro['_anio'].dropna().unique())}")

    if args.extraer or hacer_todo:
        extraer(set(filtro["Folio"]))
    if args.generar or hacer_todo:
        generar(filtro)


if __name__ == "__main__":
    main()
